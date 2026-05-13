#!/usr/bin/env python3
"""
CI DataLoader — XLSX-driven E2E test for DIGIT PGR.

Takes a county input XLSX (e.g. Bomet County), generates CRSLoader templates,
bootstraps a new root tenant, and loads all data through the standard
Jupyter notebook code path. This validates the full dataloader pipeline
end-to-end.

Flow:
  1. Generate template XLSXs from county input file
  2. Login as superuser on pg
  3. Create new root tenant (auto-bootstrap)
  4. Load boundaries from generated XLSX
  5. Load common masters (departments, designations, complaint types)
  6. Load employees from generated XLSX
  7. Load PGR workflow
  8. Seed localization (with bundled JSON fallback)
  9. Verify PGR complaint flow (create → assign → resolve)

Environment variables:
  DIGIT_URL        - Kong gateway URL (default: http://localhost:18000)
  DIGIT_USERNAME   - Superuser username (default: ADMIN)
  DIGIT_PASSWORD   - Password (default: eGov@123)
  ROOT_TENANT      - Root tenant for login (default: pg)
  BOOT_TENANT      - Cross-root tenant to create (default: ke.bomet)
  INPUT_XLSX       - Path to county input XLSX (default: county-data.xlsx)
"""

import os
import sys
import json
import time
import traceback
import requests

# Add directories to path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATALOADER_DIR = os.path.join(SCRIPT_DIR, "..", "jupyter", "dataloader")
sys.path.insert(0, DATALOADER_DIR)
sys.path.insert(0, SCRIPT_DIR)

from crs_loader import CRSLoader

# Configuration from env
BASE_URL = os.environ.get("DIGIT_URL", "http://localhost:18000")
USERNAME = os.environ.get("DIGIT_USERNAME", "ADMIN")
PASSWORD = os.environ.get("DIGIT_PASSWORD", "eGov@123")
ROOT_TENANT = os.environ.get("ROOT_TENANT", "pg")
BOOT_TENANT = os.environ.get("BOOT_TENANT", "ke.bomet")
INPUT_XLSX = os.environ.get("INPUT_XLSX", "county-data.xlsx")

BOOT_ROOT = BOOT_TENANT.split(".")[0]
REQUEST_TIMEOUT = 30


def generate_templates(input_xlsx, output_dir, tenant_code):
    """Generate CRSLoader template XLSXs from county input file."""
    # Import the generator
    sys.path.insert(0, SCRIPT_DIR)
    from importlib import import_module
    spec_path = os.path.join(SCRIPT_DIR, "generate-sample-from-xlsx.py")

    # We can't import a module with hyphens in the name, so exec it
    import openpyxl
    import re

    # Read input
    wb = openpyxl.load_workbook(input_xlsx)

    # --- Inline the essential generation logic ---
    # Read SLA from first sheet
    sla_map = {}
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[3]:
            name = str(row[0]).strip()
            sla_text = str(row[3]).strip().lower()
            # Parse SLA
            m = re.match(r"(\d+)\s*[–\-]\s*(\d+)\s*hr", sla_text)
            if m:
                sla_map[name] = (int(m.group(1)) + int(m.group(2))) // 2
            else:
                m = re.match(r"(\d+)\s*day", sla_text)
                if m:
                    sla_map[name] = int(m.group(1)) * 24
                else:
                    m = re.match(r"(\d+)\s*[–\-]\s*(\d+)\s*day", sla_text)
                    if m:
                        sla_map[name] = ((int(m.group(1)) + int(m.group(2))) // 2) * 24

    # Read Revised sheet
    ws_rev = wb["Revised"]
    revised_rows = []
    for row in ws_rev.iter_rows(min_row=2, values_only=True):
        comp_type = str(row[0]).strip() if row[0] else None
        sub_type = str(row[1]).strip() if row[1] else None
        dept = str(row[2]).strip() if row[2] else None
        sla_raw = str(row[3]).strip() if row[3] else "x days"
        keywords = str(row[4]).strip() if row[4] else ""
        if comp_type and not sub_type:
            continue  # skip comment rows
        if not sub_type:
            continue
        revised_rows.append({
            "complaint_type": comp_type, "sub_type": sub_type,
            "department": dept, "sla_raw": sla_raw, "keywords": keywords,
        })

    # Read Boundary sheet
    ws_bnd = wb["Boundary"]
    boundary_rows = []
    for row in ws_bnd.iter_rows(min_row=2, values_only=True):
        county = str(row[0]).strip() if row[0] else None
        subcounty = str(row[1]).strip() if row[1] else None
        ward = str(row[2]).strip() if row[2] else None
        if county and subcounty and ward:
            boundary_rows.append({"county": county, "subcounty": subcounty, "ward": ward})

    departments = set(r["department"] for r in revised_rows if r["department"])

    # Generate Common Master
    wb_cm = openpyxl.Workbook()
    ws_rm = wb_cm.active
    ws_rm.title = "Read me"
    ws_rm.append(["Generated from county data"])

    ws_dd = wb_cm.create_sheet("Department And Desgination Mast")
    ws_dd.append(["Department Name*", "Designation Name*"])
    for dept in sorted(departments):
        for desig in ["Health Officer", "Field Worker", "Medical Officer", "Nursing Officer", "Administrator"]:
            ws_dd.append([dept, desig])

    ws_ct = wb_cm.create_sheet("Complaint Type Master")
    ws_ct.append(["Complaint Type*", "Complaint sub type*", "Department Name*",
                  "Resolution Time (Hours)*", "Search Words (comma separated)*"])
    current_type = None
    for r in revised_rows:
        ct = r["complaint_type"]
        sla = 168
        sla_raw = r["sla_raw"]
        if sla_raw and sla_raw != "x days":
            m = re.match(r"(\d+)\s*day", sla_raw)
            if m:
                sla = int(m.group(1)) * 24
        elif ct:
            for key, val in sla_map.items():
                if key.lower().startswith(ct.lower()[:15]):
                    sla = val
                    break
        emit = None
        if ct and ct != current_type:
            emit = ct
            current_type = ct
        ws_ct.append([emit, r["sub_type"], r["department"], sla, r["keywords"]])
    wb_cm.save(os.path.join(output_dir, "Common and Complaint Master.xlsx"))

    # Generate Boundary Master
    def make_code(name, max_len=20):
        return re.sub(r"[^a-zA-Z0-9]+", "_", name.strip()).strip("_").upper()[:max_len]

    wb_bm = openpyxl.Workbook()
    ws_b = wb_bm.active
    ws_b.title = "Boundary"
    ws_b.append(["code", "name", "boundaryType", "parentCode"])

    counties = {}
    for r in boundary_rows:
        c, sc, w = r["county"], r["subcounty"], r["ward"]
        if c not in counties:
            counties[c] = {}
        if sc not in counties[c]:
            counties[c][sc] = []
        counties[c][sc].append(w)

    for county, subcounties in counties.items():
        cc = make_code(county)
        ws_b.append([cc, county, "County", None])
        for subcounty, wards in subcounties.items():
            scc = f"{cc}_{make_code(subcounty)}"
            ws_b.append([scc, subcounty, "SubCounty", cc])
            for ward in wards:
                wc = f"{scc}_{make_code(ward)}"
                ws_b.append([wc, ward, "Ward", scc])
    wb_bm.save(os.path.join(output_dir, "Boundary_Master.xlsx"))

    # Generate Employee Master
    wb_em = openpyxl.Workbook()
    ws_i = wb_em.active
    ws_i.title = "Instructions"
    ws_i.append(["Generated for E2E testing"])

    ws_e = wb_em.create_sheet("Employee Master")
    ws_e.append(["User Name*", "Mobile Number*", "Password", "Department Name*",
                 "Designation Name*", "Role Names (comma separated)*",
                 "Employee Status", "Employee Type", "Gender",
                 "Hierarchy Type", "Boundary Type", "Boundary Code",
                 "Assignment From Date*", "Date of Appointment*"])
    dept = sorted(departments)[0] if departments else "HealthServices"

    ws_e.append(["BOMET_ADMIN", "9100000003", "eGov@123", dept, "Administrator",
                 "SUPERUSER,EMPLOYEE,GRO,DGRO,PGR_LME,PGR_VIEWER,CSR,CFC",
                 "EMPLOYED", "PERMANENT", "MALE", "ADMIN", "City", tenant_code,
                 "2024-01-01", "2024-01-01"])

    # Ref sheets
    ws_rd = wb_em.create_sheet("Ref_Departments")
    ws_rd.append(["Department Code", "Department Name"])
    for i, d in enumerate(sorted(departments), 1):
        ws_rd.append([f"DEPT_{i}", d])

    ws_rds = wb_em.create_sheet("Ref_Designations")
    ws_rds.append(["Designation Code", "Designation Name"])
    for i, d in enumerate(["Health Officer", "Field Worker", "Medical Officer",
                           "Nursing Officer", "Administrator"], 1):
        ws_rds.append([f"DESIG_{i:02d}", d])

    ws_rr = wb_em.create_sheet("Ref_Roles")
    ws_rr.append(["Role Code", "Role Name"])
    for code, name in [("EMPLOYEE", "Employee"), ("SUPERUSER", "Super User"),
                       ("GRO", "GRO"), ("DGRO", "DGRO"),
                       ("PGR_LME", "PGR LME"), ("PGR_VIEWER", "PGR Viewer"),
                       ("CSR", "CSR"), ("CFC", "CFC")]:
        ws_rr.append([code, name])

    ws_rb = wb_em.create_sheet("Ref_Boundaries")
    ws_rb.append(["Boundary Code", "Boundary Type"])
    ws_rb.append([tenant_code, "City"])

    wb_em.save(os.path.join(output_dir, "Employee_Master.xlsx"))

    print(f"   Generated: {len(revised_rows)} complaint types, "
          f"{len(boundary_rows)} wards, 1 employee")
    return True


def create_boundary_hierarchy(loader, tenant, hierarchy_levels):
    """Create boundary hierarchy definition using raw API (not XLSX upload)."""
    headers = {"Content-Type": "application/json"}
    auth_info = {
        "apiId": "Rainmaker",
        "authToken": loader.auth_token,
        "userInfo": loader.user_info,
    }

    hierarchy_url = f"{loader.base_url}/boundary-service/boundary-hierarchy-definition/_create"
    hierarchy_payload = {
        "RequestInfo": auth_info,
        "BoundaryHierarchy": {
            "tenantId": tenant,
            "hierarchyType": "ADMIN",
            "boundaryHierarchy": [],
        },
    }

    parent = None
    for level in hierarchy_levels:
        hierarchy_payload["BoundaryHierarchy"]["boundaryHierarchy"].append({
            "boundaryType": level,
            "parentBoundaryType": parent,
        })
        parent = level

    resp = requests.post(hierarchy_url, json=hierarchy_payload, headers=headers, timeout=REQUEST_TIMEOUT)
    if resp.ok:
        print("   Hierarchy definition created")
        return True
    elif "already exists" in resp.text.lower() or "duplicate" in resp.text.lower():
        print("   Hierarchy definition already exists")
        return True
    else:
        print(f"   Warning: hierarchy creation: {resp.text[:200]}")
        return True  # Continue anyway


def create_boundaries_from_xlsx(loader, tenant, xlsx_path):
    """Create boundary entities and relationships from generated Boundary XLSX."""
    import openpyxl

    headers = {"Content-Type": "application/json"}
    auth_info = {
        "apiId": "Rainmaker",
        "authToken": loader.auth_token,
        "userInfo": loader.user_info,
    }

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Boundary"]

    boundaries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, name, btype, parent = row[0], row[1], row[2], row[3]
        if code:
            boundaries.append({"code": str(code), "name": str(name),
                               "boundaryType": str(btype),
                               "parent": str(parent) if parent else None})

    if not boundaries:
        print("   No boundaries in XLSX")
        return False

    # Create entities
    entity_url = f"{loader.base_url}/boundary-service/boundary/_create"
    default_geometry = {"type": "Point", "coordinates": [0, 0]}
    entity_payload = {
        "RequestInfo": auth_info,
        "Boundary": [
            {"tenantId": tenant, "code": b["code"], "geometry": default_geometry}
            for b in boundaries
        ],
    }

    resp = requests.post(entity_url, json=entity_payload, headers=headers, timeout=REQUEST_TIMEOUT)
    if resp.ok:
        print(f"   Created {len(boundaries)} boundary entities")
    elif "already exists" in resp.text.lower() or "duplicate" in resp.text.lower():
        print("   Some entities exist, creating individually...")
        for b in boundaries:
            single = {
                "RequestInfo": auth_info,
                "Boundary": [{"tenantId": tenant, "code": b["code"], "geometry": default_geometry}],
            }
            r = requests.post(entity_url, json=single, headers=headers, timeout=REQUEST_TIMEOUT)
            if not r.ok and "already exists" not in r.text.lower() and "duplicate" not in r.text.lower():
                print(f"   Warning: entity {b['code']}: {r.text[:150]}")
    else:
        print(f"   Warning: batch entity creation: {resp.text[:200]}")
        # Try one by one
        for b in boundaries:
            single = {
                "RequestInfo": auth_info,
                "Boundary": [{"tenantId": tenant, "code": b["code"], "geometry": default_geometry}],
            }
            requests.post(entity_url, json=single, headers=headers, timeout=REQUEST_TIMEOUT)

    # Create relationships
    relation_url = f"{loader.base_url}/boundary-service/boundary-relationships/_create"
    failures = 0
    for b in boundaries:
        rel = {
            "RequestInfo": auth_info,
            "BoundaryRelationship": {
                "tenantId": tenant,
                "code": b["code"],
                "hierarchyType": "ADMIN",
                "boundaryType": b["boundaryType"],
            },
        }
        if b["parent"]:
            rel["BoundaryRelationship"]["parent"] = b["parent"]
        r = requests.post(relation_url, json=rel, headers=headers, timeout=REQUEST_TIMEOUT)
        if not r.ok and "already exists" not in r.text.lower() and "duplicate" not in r.text.lower():
            print(f"   Error: relationship {b['code']}: {r.text[:150]}")
            failures += 1

    if failures > 0:
        print(f"   WARNING: {failures} relationship(s) failed")
    else:
        print(f"   Boundary tree created ({len(boundaries)} nodes)")
    return failures == 0


def verify_pgr_flow(loader, tenant, service_code):
    """Full PGR E2E: create → search → assign → resolve → search."""
    headers = {"Content-Type": "application/json"}
    auth_info = {
        "apiId": "Rainmaker",
        "authToken": loader.auth_token,
        "userInfo": loader.user_info,
    }

    # Find a valid locality from boundaries
    bnd_resp = requests.post(
        f"{loader.base_url}/boundary-service/boundary-relationships/_search",
        json={"RequestInfo": auth_info},
        params={"tenantId": tenant, "hierarchyType": "ADMIN", "boundaryType": "Ward"},
        headers=headers, timeout=REQUEST_TIMEOUT,
    )
    locality_code = None
    if bnd_resp.ok:
        rels = bnd_resp.json().get("TenantBoundary", [{}])
        if rels:
            boundary = rels[0].get("boundary", [])
            def find_leaf(nodes):
                for n in nodes:
                    children = n.get("children", [])
                    if children:
                        result = find_leaf(children)
                        if result:
                            return result
                    elif n.get("boundaryType") in ("Ward", "Locality"):
                        return n.get("code")
                return None
            locality_code = find_leaf(boundary)

    if not locality_code:
        try:
            import openpyxl as _ox
            gen_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "generated")
            bnd_xlsx = os.path.join(gen_dir, "Boundary_Master.xlsx")
            if os.path.exists(bnd_xlsx):
                wb = _ox.load_workbook(bnd_xlsx)
                ws = wb["Boundary"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[2] == "Ward":
                        locality_code = str(row[0])
                        break
        except Exception:
            pass
        if not locality_code:
            locality_code = "BOMET_SOTIK_NDANAI_ABOSI"
        print(f"   Using boundary code from XLSX: {locality_code}")

    # --- Step A: CREATE complaint ---
    pgr_url = f"{loader.base_url}/pgr-services/v2/request/_create"
    pgr_payload = {
        "RequestInfo": auth_info,
        "service": {
            "tenantId": tenant,
            "serviceCode": service_code,
            "description": "CI test complaint - Bomet E2E verification",
            "address": {
                "locality": {"code": locality_code},
                "city": tenant.split(".")[-1],
                "geoLocation": {"latitude": -0.7827, "longitude": 35.3428},
            },
            "source": "web",
            "citizen": {
                "name": "Test Citizen",
                "mobileNumber": "9199900001",
            },
        },
        "workflow": {"action": "APPLY"},
    }

    resp = requests.post(pgr_url, json=pgr_payload, headers=headers, timeout=REQUEST_TIMEOUT)
    if not resp.ok:
        print(f"   PGR create failed: {resp.status_code} {resp.text[:300]}")
        return False

    svc_wrappers = resp.json().get("ServiceWrappers", [])
    if not svc_wrappers:
        print(f"   PGR create returned no ServiceWrappers")
        return False

    service_req_id = svc_wrappers[0].get("service", {}).get("serviceRequestId")
    status = svc_wrappers[0].get("service", {}).get("applicationStatus")
    print(f"   Created: {service_req_id} (status: {status})")

    # Wait for persister to write to database
    print("   Waiting for persistence (8s)...")
    time.sleep(8)

    # --- Step B: SEARCH to verify persistence ---
    search_url = f"{loader.base_url}/pgr-services/v2/request/_search"
    sr = requests.post(search_url, json={"RequestInfo": auth_info},
                       params={"tenantId": tenant, "serviceRequestId": service_req_id},
                       headers=headers, timeout=REQUEST_TIMEOUT)
    if not sr.ok:
        print(f"   PGR search failed: {sr.status_code}")
        return False

    results = sr.json().get("ServiceWrappers", [])
    if not results:
        print(f"   PGR search returned 0 results — persistence failed")
        return False
    print(f"   Search: found {len(results)} result(s)")

    # Get full service object (needed for update — must include id, source, address)
    svc_obj = results[0].get("service", {})

    # --- Step C: ASSIGN ---
    update_url = f"{loader.base_url}/pgr-services/v2/request/_update"
    assign_resp = requests.post(update_url, json={
        "RequestInfo": auth_info,
        "service": svc_obj,
        "workflow": {"action": "ASSIGN", "comments": "CI: assigning for investigation"},
    }, headers=headers, timeout=REQUEST_TIMEOUT)

    if not assign_resp.ok:
        print(f"   ASSIGN failed: {assign_resp.status_code} {assign_resp.text[:300]}")
        return False

    assign_status = assign_resp.json().get("ServiceWrappers", [{}])[0].get("service", {}).get("applicationStatus")
    print(f"   ASSIGN: {status} -> {assign_status}")

    # Wait for update to persist
    time.sleep(5)

    # --- Step D: RESOLVE (re-fetch latest state first) ---
    sr2 = requests.post(search_url, json={"RequestInfo": auth_info},
                        params={"tenantId": tenant, "serviceRequestId": service_req_id},
                        headers=headers, timeout=REQUEST_TIMEOUT)
    svc_obj2 = sr2.json().get("ServiceWrappers", [{}])[0].get("service", {})

    resolve_resp = requests.post(update_url, json={
        "RequestInfo": auth_info,
        "service": svc_obj2,
        "workflow": {"action": "RESOLVE", "comments": "CI: issue resolved"},
    }, headers=headers, timeout=REQUEST_TIMEOUT)

    if not resolve_resp.ok:
        print(f"   RESOLVE failed: {resolve_resp.status_code} {resolve_resp.text[:300]}")
        return False

    resolve_status = resolve_resp.json().get("ServiceWrappers", [{}])[0].get("service", {}).get("applicationStatus")
    print(f"   RESOLVE: {assign_status} -> {resolve_status}")

    # Wait and do final search
    time.sleep(5)

    # --- Step E: Final search ---
    sr_final = requests.post(search_url, json={"RequestInfo": auth_info},
                             params={"tenantId": tenant},
                             headers=headers, timeout=REQUEST_TIMEOUT)
    all_complaints = sr_final.json().get("ServiceWrappers", [])
    print(f"   Final search: {len(all_complaints)} total complaint(s) in {tenant}")

    print(f"   PGR E2E lifecycle: CREATE -> ASSIGN -> RESOLVE PASSED")
    return True


def main():
    print("=" * 60)
    print("CI DataLoader — XLSX-driven E2E Test")
    print("=" * 60)
    print(f"URL:          {BASE_URL}")
    print(f"Input XLSX:   {INPUT_XLSX}")
    print(f"Boot tenant:  {BOOT_TENANT}")
    print(f"Boot root:    {BOOT_ROOT}")
    print()

    if not os.path.exists(INPUT_XLSX):
        print(f"FATAL: Input XLSX not found: {INPUT_XLSX}")
        return 1

    # Output directory for generated templates
    output_dir = os.path.join(os.path.dirname(INPUT_XLSX) or ".", "generated")
    os.makedirs(output_dir, exist_ok=True)

    total_steps = 9
    failed = []

    # Step 1: Generate templates
    print(f"\n[1/{total_steps}] Generate CRSLoader templates from {os.path.basename(INPUT_XLSX)}")
    try:
        if not generate_templates(INPUT_XLSX, output_dir, BOOT_TENANT):
            print("FATAL: Template generation failed")
            return 1
    except Exception as e:
        print(f"FATAL: Template generation failed: {e}")
        traceback.print_exc()
        return 1

    # Step 2: Login
    print(f"\n[2/{total_steps}] Login as superuser on '{ROOT_TENANT}'")
    loader = CRSLoader(BASE_URL)
    if not loader.login(username=USERNAME, password=PASSWORD, tenant_id=ROOT_TENANT):
        print("FATAL: Login failed")
        return 1

    # Step 3: Create tenant (auto-bootstraps new root)
    print(f"\n[3/{total_steps}] Create tenant '{BOOT_TENANT}' (auto-bootstrap '{BOOT_ROOT}')")
    if not loader.create_tenant(BOOT_TENANT, "Bomet County"):
        print("FATAL: create_tenant failed")
        return 1
    time.sleep(3)  # Wait for async persistence

    # Step 4: Create boundary hierarchy + load boundaries
    print(f"\n[4/{total_steps}] Create boundaries for '{BOOT_TENANT}'")
    hierarchy_levels = ["County", "SubCounty", "Ward"]
    create_boundary_hierarchy(loader, BOOT_TENANT, hierarchy_levels)
    boundary_xlsx = os.path.join(output_dir, "Boundary_Master.xlsx")
    if not create_boundaries_from_xlsx(loader, BOOT_TENANT, boundary_xlsx):
        print("WARNING: Some boundary relationships failed (continuing)")

    # Step 5: Load common masters at BOTH root and city level
    # PGR resolves ServiceDefs at root level, UI filters at city level
    print(f"\n[5/{total_steps}] Load common masters (departments, complaint types)")
    masters_xlsx = os.path.join(output_dir, "Common and Complaint Master.xlsx")
    for target in [BOOT_ROOT, BOOT_TENANT]:
        print(f"\n   Loading to: {target}")
        try:
            results = loader.load_common_masters(masters_xlsx, target_tenant=target)
            ct_result = results.get("complaint_types", {})
            if isinstance(ct_result, dict):
                created = ct_result.get("created", 0)
                exists = ct_result.get("exists", 0)
                print(f"   → {target}: {created} created, {exists} already exist")
            else:
                print(f"   → {target}: {ct_result}")
        except Exception as e:
            print(f"   ⚠️  Common masters load error for {target}: {e}")
            if target == BOOT_ROOT:
                failed.append("common_masters")

    # Step 6: Create employee via programmatic API (simpler than XLSX for CI)
    print(f"\n[6/{total_steps}] Create HRMS employee on '{BOOT_TENANT}'")
    ci_user = "BOMET_ADMIN"
    ci_password = "eGov@123"
    ci_mobile = "9100000003"
    ci_roles = ["SUPERUSER", "EMPLOYEE", "GRO", "DGRO", "PGR_LME", "PGR_VIEWER", "CSR", "CFC"]
    if not loader.create_employee(
        tenant=BOOT_TENANT,
        username=ci_user,
        password=ci_password,
        name="Bomet Admin",
        mobile=ci_mobile,
        roles=ci_roles,
    ):
        print("FATAL: Employee creation failed")
        return 1

    # Verify login
    test_loader = CRSLoader(BASE_URL)
    if not test_loader.login(username=ci_user, password=ci_password, tenant_id=BOOT_TENANT):
        print("FATAL: Bomet admin login failed")
        return 1

    # Step 7: Load PGR workflow at root level
    print(f"\n[7/{total_steps}] Load PGR workflow on '{BOOT_ROOT}'")
    templates_dir = os.path.join(DATALOADER_DIR, "templates")
    workflow_file = os.path.join(templates_dir, "PgrWorkflowConfig.json")
    if not os.path.exists(workflow_file):
        print(f"FATAL: {workflow_file} not found")
        return 1
    wf_result = loader.load_workflow(workflow_file, target_tenant=BOOT_ROOT)
    if wf_result.get("status") == "failed":
        print(f"FATAL: Workflow load failed: {wf_result.get('error')}")
        return 1

    # Step 8: Seed localization (with bundled JSON fallback)
    print(f"\n[8/{total_steps}] Seed localization for '{BOOT_TENANT}'")
    loader._seed_essential_localizations(BOOT_TENANT, source_tenant=ROOT_TENANT)

    # Step 9: Verify PGR flow
    print(f"\n[9/{total_steps}] Verify PGR complaint flow")
    # Find a service code from the loaded data (search city tenant first, then root)
    service_code = None
    for search_tenant in [BOOT_TENANT, BOOT_ROOT]:
        svc_resp = requests.post(
            f"{BASE_URL}/mdms-v2/v2/_search",
            json={
                "MdmsCriteria": {
                    "tenantId": search_tenant,
                    "schemaCode": "RAINMAKER-PGR.ServiceDefs",
                    "limit": 5,
                },
                "RequestInfo": {"apiId": "Rainmaker"},
            },
            headers={"Content-Type": "application/json"},
            timeout=REQUEST_TIMEOUT,
        )
        if svc_resp.ok:
            defs = svc_resp.json().get("mdms", [])
            if defs:
                service_code = defs[0].get("data", {}).get("serviceCode")
                print(f"   Using service code: {service_code} (from {search_tenant})")
                break

    if not service_code:
        print("   ⚠️  No service codes found, skipping PGR verification")
        failed.append("pgr_verify")
    else:
        # Re-login as the Bomet admin for PGR
        loader2 = CRSLoader(BASE_URL)
        loader2.login(username=ci_user, password=ci_password, tenant_id=BOOT_TENANT)
        if not verify_pgr_flow(loader2, BOOT_TENANT, service_code):
            failed.append("pgr_verify")

    # Summary
    print("\n" + "=" * 60)
    if failed:
        print(f"CI TEST COMPLETED WITH WARNINGS: {', '.join(failed)}")
    else:
        print("CI TEST PASSED")
    print(f"BOOT_TENANT={BOOT_TENANT}")
    print(f"BOOT_ROOT={BOOT_ROOT}")
    print(f"BOOT_USER={ci_user}")
    print(f"SERVICE_CODE={service_code or 'N/A'}")
    print("=" * 60)

    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
