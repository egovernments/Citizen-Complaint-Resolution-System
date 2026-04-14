#!/usr/bin/env python3
"""
Generate CRSLoader-compatible template XLSXs from a county data XLSX.

Reads a county input file (e.g. Bomet County health complaints) and generates
the three template files that CRSLoader.load_common_masters(), load_boundaries(),
and load_employees() expect:

  1. Common and Complaint Master.xlsx  (departments, designations, complaint types)
  2. Boundary_Master.xlsx              (boundary hierarchy for boundary service)
  3. Employee_Master.xlsx              (sample employees with roles)

Usage:
  python3 generate-sample-from-xlsx.py <input.xlsx> [output_dir]

The input XLSX must have:
  - A "Revised" sheet with columns:
      Complaint Type*, Complaint sub type*, Department Name*,
      Resolution Time (Hours)*, Search Words (comma separated)
  - A "Boundary" sheet with columns:
      ADMINISTRATIVE_COUNTY, ADMINISTRATIVE_SUBCOUNTY, ADMINISTRATIVE_WARD
  - Optionally a first sheet with SLA info (24-72 hrs, 7 days, 14 days, etc.)
"""

import os
import re
import sys
import openpyxl
from openpyxl.utils import get_column_letter


def parse_sla_hours(sla_text: str) -> int:
    """Convert SLA text to hours. Returns default 168 (7 days) if unparseable."""
    if not sla_text or sla_text == "x days":
        return 168  # default 7 days

    s = str(sla_text).strip().lower()

    # "24–72 hrs" or "24-72 hrs" → average = 48
    m = re.match(r"(\d+)\s*[–\-]\s*(\d+)\s*hr", s)
    if m:
        return (int(m.group(1)) + int(m.group(2))) // 2

    # "7 days" → 168
    m = re.match(r"(\d+)\s*day", s)
    if m:
        return int(m.group(1)) * 24

    # "60-90 days" → average in hours
    m = re.match(r"(\d+)\s*[–\-]\s*(\d+)\s*day", s)
    if m:
        return ((int(m.group(1)) + int(m.group(2))) // 2) * 24

    return 168


def read_sla_from_first_sheet(wb):
    """Read SLA mapping from the first (summary) sheet."""
    ws = wb[wb.sheetnames[0]]
    sla_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[3]:
            name = str(row[0]).strip()
            sla_map[name] = parse_sla_hours(str(row[3]))
    return sla_map


def read_revised_sheet(wb):
    """Read the Revised sheet and return complaint type data."""
    ws = wb["Revised"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        comp_type = str(row[0]).strip() if row[0] else None
        sub_type = str(row[1]).strip() if row[1] else None
        dept = str(row[2]).strip() if row[2] else None
        sla_raw = str(row[3]).strip() if row[3] else "x days"
        keywords = str(row[4]).strip() if row[4] else ""

        # Skip comment rows at the bottom
        if comp_type and not sub_type:
            continue
        if not sub_type:
            continue

        rows.append({
            "complaint_type": comp_type,
            "sub_type": sub_type,
            "department": dept,
            "sla_raw": sla_raw,
            "keywords": keywords,
        })
    return rows


def read_boundary_sheet(wb):
    """Read boundary data from the Boundary sheet."""
    ws = wb["Boundary"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        county = str(row[0]).strip() if row[0] else None
        subcounty = str(row[1]).strip() if row[1] else None
        ward = str(row[2]).strip() if row[2] else None
        lat = row[3] if row[3] else None
        lon = row[4] if row[4] else None
        if county and subcounty and ward:
            rows.append({
                "county": county,
                "subcounty": subcounty,
                "ward": ward,
                "lat": lat,
                "lon": lon,
            })
    return rows


def make_code(name: str, max_len: int = 20) -> str:
    """Generate a boundary code from a name. E.g. 'Bomet East' → 'BOMET_EAST'."""
    code = re.sub(r"[^a-zA-Z0-9]+", "_", name.strip()).strip("_").upper()
    return code[:max_len]


def generate_common_master_xlsx(revised_rows, sla_map, output_path):
    """Generate Common and Complaint Master.xlsx in template format."""
    wb = openpyxl.Workbook()

    # Sheet 1: Read me
    ws_readme = wb.active
    ws_readme.title = "Read me"
    ws_readme.append(["COMMON MASTER TEMPLATE - Generated from county data"])

    # Sheet 2: Department And Desgination Mast
    ws_dept = wb.create_sheet("Department And Desgination Mast")
    ws_dept.append(["Department Name*", "Designation Name*"])

    # Collect unique departments
    departments = set()
    for row in revised_rows:
        if row["department"]:
            departments.add(row["department"])

    # Standard designations for health services
    designations = [
        "Health Officer",
        "Field Worker",
        "Medical Officer",
        "Nursing Officer",
        "Administrator",
    ]

    for dept in sorted(departments):
        for desig in designations:
            ws_dept.append([dept, desig])

    # Sheet 3: Complaint Type Master
    ws_ct = wb.create_sheet("Complaint Type Master")
    ws_ct.append([
        "Complaint Type*",
        "Complaint sub type*",
        "Department Name*",
        "Resolution Time (Hours)*",
        "Search Words (comma separated)*",
    ])

    current_type = None
    for row in revised_rows:
        comp_type = row["complaint_type"]

        # Resolve SLA
        sla_hours = parse_sla_hours(row["sla_raw"])
        if sla_hours == 168 and comp_type:  # default — try SLA map
            for key, val in sla_map.items():
                if comp_type and key.lower().startswith(comp_type.lower()[:15]):
                    sla_hours = val
                    break

        # Only emit complaint type in first row of each group
        emit_type = None
        if comp_type and comp_type != current_type:
            emit_type = comp_type
            current_type = comp_type

        ws_ct.append([
            emit_type,
            row["sub_type"],
            row["department"],
            sla_hours,
            row["keywords"],
        ])

    wb.save(output_path)
    print(f"  Created: {output_path} ({len(revised_rows)} complaint sub-types)")


def generate_boundary_xlsx(boundary_rows, output_path):
    """Generate Boundary_Master.xlsx in the format CRSLoader.load_boundaries() expects.

    Format: code | name | boundaryType | parentCode
    Hierarchy: County → SubCounty → Ward
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Boundary"
    ws.append(["code", "name", "boundaryType", "parentCode"])

    # Build hierarchy
    counties = {}
    for row in boundary_rows:
        c = row["county"]
        sc = row["subcounty"]
        w = row["ward"]
        if c not in counties:
            counties[c] = {}
        if sc not in counties[c]:
            counties[c][sc] = []
        counties[c][sc].append(w)

    total = 0
    for county, subcounties in counties.items():
        county_code = make_code(county)
        ws.append([county_code, county, "County", ""])
        total += 1

        for subcounty, wards in subcounties.items():
            sc_code = f"{county_code}_{make_code(subcounty)}"
            ws.append([sc_code, subcounty, "SubCounty", county_code])
            total += 1

            for ward in wards:
                w_code = f"{sc_code}_{make_code(ward)}"
                ws.append([w_code, ward, "Ward", sc_code])
                total += 1

    wb.save(output_path)
    print(f"  Created: {output_path} ({total} boundaries)")


def generate_employee_xlsx(departments, tenant_code, output_path):
    """Generate Employee_Master.xlsx with sample employees."""
    wb = openpyxl.Workbook()

    # Instructions sheet
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr.append(["Employee Master - Generated for E2E testing"])

    # Employee Master sheet
    ws_emp = wb.create_sheet("Employee Master")
    ws_emp.append([
        "User Name*", "Mobile Number*", "Password",
        "Department Name*", "Designation Name*",
        "Role Names (comma separated)*",
        "Employee Status", "Employee Type", "Gender",
        "Hierarchy Type", "Boundary Type", "Boundary Code",
        "Assignment From Date*", "Date of Appointment*",
    ])

    dept = sorted(departments)[0] if departments else "HealthServices"

    # GRO employee
    ws_emp.append([
        "BOMET_GRO", "9100000001", "eGov@123",
        dept, "Health Officer",
        "EMPLOYEE,GRO,DGRO,PGR_VIEWER",
        "EMPLOYED", "PERMANENT", "MALE",
        "ADMIN", "City", tenant_code,
        "2024-01-01", "2024-01-01",
    ])

    # LME employee
    ws_emp.append([
        "BOMET_LME", "9100000002", "eGov@123",
        dept, "Field Worker",
        "EMPLOYEE,PGR_LME,PGR_VIEWER",
        "EMPLOYED", "PERMANENT", "FEMALE",
        "ADMIN", "City", tenant_code,
        "2024-01-01", "2024-01-01",
    ])

    # Admin employee
    ws_emp.append([
        "BOMET_ADMIN", "9100000003", "eGov@123",
        dept, "Administrator",
        "SUPERUSER,EMPLOYEE,GRO,DGRO,PGR_LME,PGR_VIEWER,CSR,CFC",
        "EMPLOYED", "PERMANENT", "MALE",
        "ADMIN", "City", tenant_code,
        "2024-01-01", "2024-01-01",
    ])

    # Ref sheets for dropdowns
    ws_ref_dept = wb.create_sheet("Ref_Departments")
    ws_ref_dept.append(["Department Code", "Department Name"])
    for i, d in enumerate(sorted(departments), 1):
        ws_ref_dept.append([f"DEPT_{i}", d])

    ws_ref_desig = wb.create_sheet("Ref_Designations")
    ws_ref_desig.append(["Designation Code", "Designation Name"])
    for i, d in enumerate(["Health Officer", "Field Worker", "Medical Officer",
                            "Nursing Officer", "Administrator"], 1):
        ws_ref_desig.append([f"DESIG_{i:02d}", d])

    ws_ref_roles = wb.create_sheet("Ref_Roles")
    ws_ref_roles.append(["Role Code", "Role Name"])
    for code, name in [
        ("EMPLOYEE", "Employee"), ("SUPERUSER", "Super User"),
        ("GRO", "Grievance Routing Officer"), ("DGRO", "Deputy GRO"),
        ("PGR_LME", "PGR Last Mile Employee"), ("PGR_VIEWER", "PGR Viewer"),
        ("CSR", "Customer Support Rep"), ("CFC", "Call Center Agent"),
    ]:
        ws_ref_roles.append([code, name])

    ws_ref_bnd = wb.create_sheet("Ref_Boundaries")
    ws_ref_bnd.append(["Boundary Code", "Boundary Type"])
    ws_ref_bnd.append([tenant_code, "City"])

    wb.save(output_path)
    print(f"  Created: {output_path} (3 employees)")


def main():
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <input.xlsx> [output_dir] [tenant_code]")
        return 1

    input_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.dirname(input_path) or "."
    tenant_code = sys.argv[3] if len(sys.argv) > 3 else "ke.bomet"

    if not os.path.exists(input_path):
        print(f"Error: {input_path} not found")
        return 1

    os.makedirs(output_dir, exist_ok=True)

    print(f"Input:  {input_path}")
    print(f"Output: {output_dir}")
    print(f"Tenant: {tenant_code}")
    print()

    wb = openpyxl.load_workbook(input_path)

    # Read source data
    sla_map = read_sla_from_first_sheet(wb)
    revised_rows = read_revised_sheet(wb)
    boundary_rows = read_boundary_sheet(wb)

    # Collect departments
    departments = set()
    for row in revised_rows:
        if row["department"]:
            departments.add(row["department"])

    print(f"Source data: {len(revised_rows)} complaint sub-types, {len(boundary_rows)} wards, {len(sla_map)} SLA entries")
    print()

    # Generate template XLSXs
    generate_common_master_xlsx(
        revised_rows, sla_map,
        os.path.join(output_dir, "Common and Complaint Master.xlsx")
    )

    generate_boundary_xlsx(
        boundary_rows,
        os.path.join(output_dir, "Boundary_Master.xlsx")
    )

    generate_employee_xlsx(
        departments, tenant_code,
        os.path.join(output_dir, "Employee_Master.xlsx")
    )

    print()
    print("Done! Generated files are ready for CRSLoader.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
