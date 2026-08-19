#!/usr/bin/env python3
"""Generate filled-in example spreadsheets for the onboarding wizard.

The configurator has a "Download Template" button in every phase, and that
button is always the authority on the columns. This script produces the same
workbooks with realistic data already in them, so you can see what a finished
sheet looks like before filling in your own — and so a small deployment can
be onboarded by editing these rather than starting from an empty template.

    python3 make-onboarding-examples.py --out-dir ./example-onboarding \\
        --root kenya --city nairobi

Writes five files, in upload order:

    01-tenant-and-branding.xlsx        Phase 1
    02-boundaries.xlsx                 Phase 2 (the Upload from Excel path)
    03-departments-and-designations.xlsx   Phase 3, step 3.1
    04-complaint-hierarchy.xlsx        Phase 3, step 3.2
    05-employees.xlsx                  Phase 4

Sheet names matter: the parser looks for them by name and falls back to the
first sheet only for the tenant file.

Column headers matter more than they look, and the rule is not uniform:

  * Tenant sheet only — matched leniently (case-insensitive, `*` and extra
    whitespace ignored, substring accepted), so "Tenant Code*", "Tenant Code"
    and "tenantCode" are the same column.
  * Boundary / Department / Designation / Employee — each field has a fixed
    list of accepted spellings (e.g. `code`, `Code`, `boundaryCode`). Anything
    outside that list is not seen.
  * Complaint hierarchy level columns — matched **exactly** against the level
    codes you typed into the wizard, with the single concession that `_` may
    be written as a space. A renamed level column silently reads as empty and
    the row is rejected as "missing <level>".

This script emits the exact strings each parser expects, so the generated
files import as-is. Edit a header by hand and you are on your own.

Needs openpyxl:  pip install openpyxl
"""

from __future__ import annotations

import argparse
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is not installed. In a virtualenv: pip install openpyxl")


# The complaint-hierarchy template is generated from whatever levels the
# operator typed into the wizard, so these are only the default shape.
DEFAULT_COMPLAINT_LEVELS = ["Category", "Sub Category", "Complaint Type"]


def write_sheet(ws, header, rows):
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(list(row))
    for i, name in enumerate(header, start=1):
        width = max(len(str(name)) + 2, *(len(str(r[i - 1])) + 2 for r in rows)) if rows else len(name) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(max(width, 14), 42)
    ws.freeze_panes = "A2"


def tenant_workbook(root, city, city_title, lat, lng, district):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tenant Info"
    write_sheet(
        ws,
        ["Tenant Display Name*", "Tenant Code*", "Tenant Type*", "Logo File Path*",
         "Latitude", "Longitude", "City Name", "District Name"],
        [[f"{city_title} City Council", f"{root}.{city}", "City", "",
          lat, lng, city_title, district]],
    )
    # Header only — a row of empty strings reads as a data row with four blank
    # fields, not as "nothing to import". Kept for backwards compatibility; the
    # recommended path is to upload the images in step 1.2, which fills these in
    # with filestore ids for you.
    write_sheet(
        wb.create_sheet("Tenant Branding Details"),
        ["Banner URL", "Logo URL", "Logo URL (White)", "State Logo"],
        [],
    )
    return wb


def boundary_workbook(city_title, lat, lng):
    # Parents before children, and every parentCode must already appear in the
    # code column above it. The top row has no parent.
    rows = [
        ["COUNTY_001", city_title, "County", "", lat, lng],
        ["SUBCOUNTY_001", f"{city_title} Central", "Sub-County", "COUNTY_001", "", ""],
        ["SUBCOUNTY_002", f"{city_title} East", "Sub-County", "COUNTY_001", "", ""],
        ["WARD_001", "Market Ward", "Ward", "SUBCOUNTY_001", "", ""],
        ["WARD_002", "Riverside Ward", "Ward", "SUBCOUNTY_001", "", ""],
        ["WARD_003", "Industrial Ward", "Ward", "SUBCOUNTY_002", "", ""],
        ["WARD_004", "Hillside Ward", "Ward", "SUBCOUNTY_002", "", ""],
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Boundary"
    write_sheet(ws, ["code", "name", "boundaryType", "parentCode", "latitude", "longitude"], rows)
    return wb


def masters_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Department"
    write_sheet(
        ws,
        ["code", "name", "active"],
        [["WATER", "Water and Sanitation", "true"],
         ["ROADS", "Roads and Transport", "true"],
         ["WASTE", "Solid Waste Management", "true"],
         ["HEALTH", "Public Health", "true"]],
    )
    write_sheet(
        wb.create_sheet("Designation"),
        ["code", "name", "description", "department", "active"],
        [["OFFICER", "Field Officer", "Handles complaints on the ground", "WATER,ROADS,WASTE,HEALTH", "true"],
         ["SUPERVISOR", "Supervisor", "Assigns work and reviews resolutions", "WATER,ROADS,WASTE,HEALTH", "true"],
         ["DIRECTOR", "Director", "Department head", "WATER,ROADS,WASTE,HEALTH", "true"]],
    )
    return wb


def complaint_hierarchy_workbook(levels):
    # One row per complaint type, carrying its full path. Rows that share the
    # earlier columns fold into one branch of the citizen menu.
    leaf_extras = ["Department Name*", "Resolution Time (Hours)*", "Search Words*"]
    rows = [
        ["Water", "Supply", "No water supply", "WATER", 24, "no water, dry tap, supply"],
        ["Water", "Supply", "Low water pressure", "WATER", 48, "low pressure, weak flow"],
        ["Water", "Leaks", "Burst pipe", "WATER", 8, "burst, leak, pipe"],
        ["Water", "Leaks", "Blocked drain", "WATER", 24, "drain, blocked, sewer"],
        ["Roads", "Surface", "Pothole", "ROADS", 72, "pothole, road damage"],
        ["Roads", "Lighting", "Street light not working", "ROADS", 48, "street light, lamp, dark"],
        ["Waste", "Collection", "Missed waste collection", "WASTE", 24, "garbage, bin, collection"],
        ["Waste", "Dumping", "Illegal dumping", "WASTE", 48, "dumping, fly tipping"],
        ["Health", "Sanitation", "Public toilet out of order", "HEALTH", 48, "toilet, sanitation"],
        # One complaint type handled by two departments: the first is primary.
        ["Health", "Pests", "Mosquito breeding site", "HEALTH,WATER", 72, "mosquito, stagnant water"],
    ]
    # The rows above are authored for a three-level hierarchy. Reshape only the
    # PATH part to however many levels the operator defined, and keep the three
    # fixed leaf columns intact — they are positional, so slicing the whole row
    # slides the department into a level column and the SLA into the department.
    n = len(levels)
    if n != 3:
        reshaped = []
        for row in rows:
            path, fixed = list(row[:3]), list(row[3:])
            if n < 3:
                # Keep the outermost group and the leaf; drop the middle levels.
                path = [path[0]] + path[-(n - 1):]
            else:
                # Repeat the sub-category to fill the extra levels. The leaf and
                # its immediate parent stay put, so the generated serviceCode
                # (parent + leaf) is unchanged and still unique per row.
                path = path[:2] + [path[1]] * (n - 3) + path[2:]
            assert len(path) == n, (n, path)
            reshaped.append(path + fixed)
        rows = reshaped
    wb = Workbook()
    ws = wb.active
    ws.title = "ComplaintHierarchy"
    write_sheet(ws, list(levels) + leaf_extras, rows)
    return wb


def employee_workbook(mobile_prefix):
    def mob(n):
        return f"{mobile_prefix}{n:03d}"
    rows = [
        ["EMP001", "Amina Otieno", "amina.otieno", mob(1), "amina@example.com", "FEMALE",
         "1988-04-12", "WATER", "SUPERVISOR", "EMPLOYEE,GRO,CSR", "COUNTY_001", "2024-01-15"],
        ["EMP002", "Brian Kimani", "brian.kimani", mob(2), "brian@example.com", "MALE",
         "1990-09-03", "WATER", "OFFICER", "EMPLOYEE,PGR_LME", "WARD_001", "2024-02-01"],
        ["EMP003", "Grace Wanjiru", "grace.wanjiru", mob(3), "grace@example.com", "FEMALE",
         "1985-12-20", "ROADS", "SUPERVISOR", "EMPLOYEE,DGRO", "SUBCOUNTY_001", "2023-11-06"],
        ["EMP004", "Daniel Mwangi", "daniel.mwangi", mob(4), "daniel@example.com", "MALE",
         "1992-06-30", "ROADS,WASTE", "OFFICER", "EMPLOYEE,PGR_LME", "WARD_003", "2024-03-11"],
        ["EMP005", "Faith Njeri", "faith.njeri", mob(5), "faith@example.com", "FEMALE",
         "1994-02-08", "WASTE", "OFFICER", "EMPLOYEE,PGR_LME", "WARD_002", "2024-05-20"],
        ["EMP006", "Peter Ochieng", "peter.ochieng", mob(6), "peter@example.com", "MALE",
         "1980-07-17", "HEALTH", "DIRECTOR", "EMPLOYEE,GRO,PGR_VIEWER", "COUNTY_001", "2022-08-01"],
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee"
    write_sheet(
        ws,
        ["employeeCode", "name", "userName", "mobileNumber", "emailId", "gender", "dob",
         "department", "designation", "roles", "jurisdictions", "dateOfAppointment"],
        rows,
    )
    return wb


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--out-dir", default="example-onboarding",
                    help="directory to write the workbooks into (created if needed)")
    ap.add_argument("--root", default="mycountry", help="root tenant code, e.g. kenya")
    ap.add_argument("--city", default="mycity", help="city segment, e.g. nairobi")
    ap.add_argument("--city-name", default=None, help="display name (default: --city, title-cased)")
    ap.add_argument("--district", default="Central", help="district name for the tenant sheet")
    ap.add_argument("--lat", default="-1.2864", help="city latitude")
    ap.add_argument("--lng", default="36.8172", help="city longitude")
    ap.add_argument("--mobile-prefix", default="712345",
                    help="leading digits for the example phone numbers; they must satisfy "
                         "the tenant's mobileNumberRegex or the employee upload is rejected")
    ap.add_argument("--complaint-levels", default=",".join(DEFAULT_COMPLAINT_LEVELS),
                    help="comma-separated complaint-hierarchy levels, top to leaf; must match "
                         "what you typed into the wizard in step 3.2")
    args = ap.parse_args()

    city_title = args.city_name or args.city.replace("-", " ").replace("_", " ").title()
    levels = [s.strip() for s in args.complaint_levels.split(",") if s.strip()]
    if len(levels) < 2:
        sys.exit("--complaint-levels needs at least two levels (a group and a leaf)")

    os.makedirs(args.out_dir, exist_ok=True)
    files = [
        ("01-tenant-and-branding.xlsx",
         tenant_workbook(args.root, args.city, city_title, args.lat, args.lng, args.district)),
        ("02-boundaries.xlsx", boundary_workbook(city_title, args.lat, args.lng)),
        ("03-departments-and-designations.xlsx", masters_workbook()),
        ("04-complaint-hierarchy.xlsx", complaint_hierarchy_workbook(levels)),
        ("05-employees.xlsx", employee_workbook(args.mobile_prefix)),
    ]
    for name, wb in files:
        path = os.path.join(args.out_dir, name)
        wb.save(path)
        print(f"wrote {path}")

    print(f"""
Five example workbooks are in {args.out_dir}/, in upload order.

Before uploading them to a real deployment, check three things:
  1. the phone numbers match your tenant's mobile rule (core_mobile_configs)
  2. the boundary level names match what you typed in the wizard's Phase 2
  3. the complaint-hierarchy column headers match the levels you defined in
     step 3.2 — the template is generated from them, so they have to agree
""")


if __name__ == "__main__":
    main()
