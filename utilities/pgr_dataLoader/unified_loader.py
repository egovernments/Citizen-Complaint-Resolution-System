"""
Unified PGR Data Loader - Core Module
All code for reading Excel and generating API payloads
Users should not modify this file directly
"""

import pandas as pd
import json
import math
import warnings
import requests
import time
from typing import Dict, List, Any
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def clean_nans(obj):
    """
    Recursively clean NaN values and convert dates for JSON serialization
    """
    if isinstance(obj, dict):
        cleaned = {}
        for k, v in obj.items():
            cleaned[k] = clean_nans(v)
        return cleaned
    elif isinstance(obj, list):
        return [clean_nans(item) for item in obj]
    elif isinstance(obj, float) and math.isnan(obj):
        return None  # Convert NaN to null
    elif isinstance(obj, (pd.Timestamp, datetime)):
        return obj.isoformat()  # Convert Timestamp/datetime to ISO string
    elif pd.isna(obj):
        return None  # Handle any other pandas NA types
    else:
        return obj  # Keep None as None (will be null in JSON)


# ============================================================================
# EXCEL READER CLASS
# ============================================================================

class UnifiedExcelReader:
    """Read unified Excel template and generate API payloads"""

    def __init__(self, excel_file: str):
        self.excel_file = excel_file

    # ========================================================================
    # TENANT MASTER READERS (PHASE 1)
    # ========================================================================

    def read_tenant_info(self):
        """
        Reads the NEW Tenant Master template.
        Updated to match:
        - Tenant Display Name*
        - Tenant Code*
        - Tenant Type*
        - Logo File Path*
        - Latitude, Longitude
        - City Name
        - District Name
        - Address
        - Tenant Website
        """

        df = pd.read_excel(self.excel_file, sheet_name='Tenant Info')

        tenants = []
        localizations = []
        district_counter = {}  # For auto-generating district codes
        city_counter = {}   

        for _, row in df.iterrows():
            
            # Skip empty rows
            if pd.isna(row.get('Tenant Display Name*')) or pd.isna(row.get('Tenant Code*\n(To be filled by ADMIN)')):
                continue

            tenant_name = str(row['Tenant Display Name*']).strip()
            tenant_code_col = 'Tenant Code*\n(To be filled by ADMIN)'
            tenant_code = str(row[tenant_code_col]).strip().lower()

            tenant_type = str(row.get('Tenant Type*', '')).strip()
            logo_path = str(row.get('Logo File Path*', '')).strip()

            city_name = str(row.get('City Name', '')).strip()
            district_name = str(row.get('District Name', '')).strip()
            address = str(row.get('Address', '')).strip()
            website = str(row.get('Tenant Website', '')).strip()

            latitude = float(row['Latitude']) if pd.notna(row.get('Latitude')) else 0.0
            longitude = float(row['Longitude']) if pd.notna(row.get('Longitude')) else 0.0

            # Build city object
            district_code = ""
            if district_name:
                if district_name not in district_counter:
                    district_counter[district_name] = len(district_counter) + 1
                district_code = f"District_{district_counter[district_name]:03d}"
            city_code = ""
            if city_name:
                if city_name not in city_counter:
                    city_counter[city_name] = len(city_counter) + 1
                city_code = f"City_{city_counter[city_name]:03d}"
    
            city = {
               'code': city_code,
                'name': city_name or tenant_name,
                'districtName': district_name,
                'districtTenantCode': district_code,
                'ulbGrade':'',
                'latitude': latitude,
                'longitude': longitude,
                'localName': tenant_name,
                'captcha': 'true'
            }

            # Tenant object
            tenant = {
                'code': tenant_code,
                'name': tenant_name,
                'type': tenant_type,
                'emailId': "",
                'contactNumber': "",
                'address': address,
                'domainUrl': website or "https://example.com",
                'logoId': logo_path,
                'imageId': logo_path,
                'description': tenant_name,
                'twitterUrl': "",
                'facebookUrl': "",
                'OfficeTimings': {'Mon - Fri': '10:00 AM - 5:00 PM'},
                'city': city
            }

            tenants.append(tenant)

            # Add localization
            loc_code = f"TENANT_TENANTS_{tenant_code.upper().replace('.', '_')}"
            localizations.append({
                "code": loc_code,
                "message": tenant_name,
                "module": "rainmaker-common",
                "locale": "en_IN"
            })
        return tenants, localizations


    def read_tenant_branding(self, tenant_id: str):
        """Read Tenant Branding sheet

        Returns:
            list: Branding information records for MDMS upload
        """
        try:
            df = pd.read_excel(self.excel_file, sheet_name='Tenant Branding Deatils')
        except Exception as e:
            print(f"⚠️ Could not read 'Tenant Branding Deatils' sheet: {str(e)}")
            return []

        branding_list = []
        for _, row in df.iterrows():
                branding_record = {
                     'code':tenant_id,
                    'name':tenant_id,
                    'bannerUrl': str(row.get('Banner URL', '')) if pd.notna(row.get('Banner URL')) else "",
                    'logoUrl': str(row.get('Logo URL', '')) if pd.notna(row.get('Logo URL')) else "",
                    'logoUrlWhite': str(row.get('Logo URL (White)', '')) if pd.notna(row.get('Logo URL (White)')) else "",
                    'statelogo': str(row.get('State Logo', '')) if pd.notna(row.get('State Logo')) else ""
                }
                branding_list.append(branding_record)

        return branding_list

    # ========================================================================
    # COMMON MASTER READERS (PHASE 2)
    # ========================================================================

    def read_departments_designations(self, tenant_id: str):
        """Read combined Department and Designation sheet from Common Master Excel
        Auto-generates codes for departments and designations

        Args:
            tenant_id: Tenant ID for localization context

        Returns:
            tuple: (departments_list, designations_list, dept_localization, desig_localization, dept_name_to_code_mapping)
        """
        df = pd.read_excel(self.excel_file, sheet_name='Department And Desgination Mast')

        departments = []
        designations = []
        dept_localizations = []
        desig_localizations = []

        dept_counter = {}
        dept_name_to_code = {}  # Mapping for complaint types
        desig_counter = 1

        for _, row in df.iterrows():
            dept_name = row.get('Department Name*')
            desig_name = row.get('Designation Name*')

            # Skip empty rows
            if pd.isna(dept_name) or str(dept_name).strip() == '':
                continue

            dept_name = str(dept_name).strip()

            # Auto-generate department code
            if dept_name not in dept_counter:
                dept_counter[dept_name] = len(dept_counter) + 1
                dept_code = f"DEPT_{dept_counter[dept_name]}"

                # Store mapping from name to code
                dept_name_to_code[dept_name] = dept_code

                # Add department
                departments.append({
                    'code': dept_code,
                    'name': dept_name,
                    'active': True
                })

                # Auto-generate department localization
                loc_code = f"COMMON_MASTERS_DEPARTMENT_{dept_code}"
                dept_localizations.append({
                    'code': loc_code,
                    'message': dept_name,
                    'module': 'rainmaker-common',
                    'locale': 'en_IN'
                })
            else:
                dept_code = f"DEPT_{dept_counter[dept_name]}"

            # Add designation if present
            if pd.notna(desig_name) and str(desig_name).strip() != '':
                desig_name = str(desig_name).strip()
                desig_code = f"DESIG_{desig_counter:02d}"
                desig_counter += 1

                designations.append({
                    'code': desig_code,
                    'name': desig_name,
                    'department': [dept_code],
                    'active': True,
                    'description': f'{desig_name} - {dept_name}'
                })

                # Auto-generate designation localization
                loc_code = f"COMMON_MASTERS_{desig_code}"
                desig_localizations.append({
                    'code': loc_code,
                    'message': desig_name,
                    'module': 'rainmaker-common',
                    'locale': 'en_IN'
                })

        return departments, designations, dept_localizations, desig_localizations, dept_name_to_code

    def read_complaint_types(self, tenant_id: str, dept_name_to_code: dict = None):
        """Read Complaint Type Master from Common Master Excel
        Auto-generates service codes and handles hierarchical structure

        Args:
            tenant_id: Tenant ID for context
            dept_name_to_code: Dictionary mapping department names to codes

        Returns:
            tuple: (complaint_types_list, localization_list)
        """
        df = pd.read_excel(self.excel_file, sheet_name='Complaint Type Master')

        complaint_types = []
        localizations = []
        current_parent = None
        localized_parent_types = set()

        # If no mapping provided, create empty dict
        if dept_name_to_code is None:
            dept_name_to_code = {}

        for _, row in df.iterrows():
            # Check if this is a parent row (has Complaint Type* filled)
            if pd.notna(row.get('Complaint Type*')):
                parent_type = str(row['Complaint Type*']).strip()

                # Get department name and convert to code
                dept_name = str(row['Department Name*']).strip() if pd.notna(row.get('Department Name*')) else None
                dept_code = dept_name_to_code.get(dept_name, dept_name) if dept_name else None

                current_parent = {
                    'type': parent_type,
                    'department': dept_code,
                    'slaHours': int(float(row['Resolution Time (Hours)*'])) if pd.notna(row.get('Resolution Time (Hours)*')) else None,
                    'keywords': str(row['Search Words (comma separated)*']).strip() if pd.notna(row.get('Search Words (comma separated)*')) else None,
                    'order': int(float(row['Priority'])) if pd.notna(row.get('Priority')) else None
                }

                # Auto-generate localization for parent type (only once)
                if parent_type not in localized_parent_types:
                    parent_type_code = ''.join(word.capitalize() for word in parent_type.split())
                    loc_code = f"SERVICEDFS.{parent_type_code.upper()}"
                    localizations.append({
                        'code': loc_code,
                        'message': parent_type,
                        'module': 'rainmaker-pgr',
                        'locale': 'en_IN'
                    })
                    localized_parent_types.add(parent_type)

            # Every row should have a sub-type
            if pd.notna(row.get('Complaint sub type*')) and str(row['Complaint sub type*']).strip() != '':
                sub_type_name = str(row['Complaint sub type*']).strip()

                # Auto-generate service code from sub-type name
                service_code = ''.join(word.capitalize() for word in sub_type_name.split())

                menu_path_value = current_parent['type'] if current_parent else sub_type_name

                ct = {
                    'serviceCode': service_code,
                    'name': sub_type_name,
                    'menuPath': menu_path_value,
                    'menuPathName': menu_path_value,
                    'active': True
                }

                # Add parent-level fields
                if current_parent:
                    if current_parent.get('department'):
                        ct['department'] = current_parent['department']
                    if current_parent.get('slaHours'):
                        ct['slaHours'] = current_parent['slaHours']
                    if current_parent.get('keywords'):
                        ct['keywords'] = current_parent['keywords']
                    if current_parent.get('order'):
                        ct['order'] = current_parent['order']

                complaint_types.append(ct)

                # Auto-generate localization for sub-type
                loc_code = f"SERVICEDFS.{service_code.upper()}"
                localizations.append({
                    'code': loc_code,
                    'message': sub_type_name,
                    'module': 'rainmaker-pgr',
                    'locale': 'en_IN'
                })

        return complaint_types, localizations

    # ========================================================================
    # OTHER UTILITY READERS (Employee, Boundary, Workflow, Localization)
    # ========================================================================

    def read_employees(self):
        """Read employee data from Employee sheet (with embedded jurisdiction data)"""
        df = pd.read_excel(self.excel_file, sheet_name='Employee')

        employees = []
        for _, row in df.iterrows():
            emp = {
                'code': row['Employee Code'],
                'tenantId': row['Tenant ID'],
                'employeeStatus': row['Employee Status'],
                'employeeType': row['Employee Type'],
                'dateOfAppointment': row['Date of Appointment'],
                'assignments': [{
                    'fromDate': row['Assignment From Date'],
                    'toDate': row['Assignment To Date'] if pd.notna(row['Assignment To Date']) else None,
                    'isCurrentAssignment': str(row['Is Current Assignment']).upper() == 'TRUE',
                    'department': row['Department Code'],
                    'designation': row['Designation Code']
                }],
                'user': {
                    'mobileNumber': row['User Mobile Number'],
                    'name': row['User Name'],
                    'emailId': row['User Email'] if pd.notna(row['User Email']) else '',
                    'gender': row['User Gender'],
                    'dob': row['User Date of Birth'],
                    'correspondenceAddress': row['User Correspondence Address']
                }
            }

            # Read embedded jurisdiction data
            jurisdiction_roles = []
            if pd.notna(row.get('Jurisdiction Roles (comma separated)')):
                role_codes = [r.strip() for r in str(row['Jurisdiction Roles (comma separated)']).split(',')]
                jurisdiction_roles = [{'code': code, 'name': code, 'tenantId': row['Jurisdiction Tenant ID']} for code in role_codes]

            # Read user roles
            user_roles = []
            if pd.notna(row.get('User Roles (comma separated)')):
                user_role_codes = [r.strip() for r in str(row['User Roles (comma separated)']).split(',')]
                user_roles = [{'code': code, 'name': code, 'tenantId': row['Tenant ID']} for code in user_role_codes]

            emp['user']['roles'] = user_roles

            emp['jurisdictions'] = [{
                'hierarchy': row['Jurisdiction Hierarchy'],
                'boundaryType': row['Jurisdiction Boundary Type'],
                'boundary': row['Jurisdiction Boundary Code'],
                'tenantId': row['Jurisdiction Tenant ID'],
                'roles': jurisdiction_roles
            }]

            # Add optional fields
            emp['serviceHistory'] = []
            emp['education'] = []
            emp['tests'] = []

            if pd.notna(row.get('Service History')):
                emp['serviceHistory'] = [row['Service History']]
            if pd.notna(row.get('Education')):
                emp['education'] = [row['Education']]
            if pd.notna(row.get('Tests')):
                emp['tests'] = [row['Tests']]

            employees.append(emp)

        return employees

    def read_boundary_hierarchy(self):
        """Read boundary hierarchy definition"""
        df = pd.read_excel(self.excel_file, sheet_name='Hierarchy_Definition')

        if len(df) == 0:
            return None

        row = df.iloc[0]
        hierarchy = {
            'tenantId': row['City Code'],
            'hierarchyType': row['Hierarchy Type'],
            'boundaryHierarchy': []
        }
        prev_level = None

        for col in df.columns:
            if col.startswith('Level'):
                level_name = row[col]
                if pd.notna(level_name):
                    hierarchy['boundaryHierarchy'].append({
                        'boundaryType': str(level_name),
                        'parentBoundaryType': prev_level,
                        'active': True
                    })
                    prev_level = str(level_name)

        return hierarchy

    def read_boundary_entities(self):
        """Read boundary entities with GeoJSON"""
        df = pd.read_excel(self.excel_file, sheet_name='Boundary_Entities')

        boundaries = []
        for _, row in df.iterrows():
            # Read coordinates - column is 'Polygon Coordinates (JSON)'
            coord_string = row['Polygon Coordinates (JSON)']

            if pd.notna(coord_string):
                try:
                    coordinates_array = json.loads(coord_string)
                    coordinates = [coordinates_array]
                except:
                    coordinates = [[[]]]
            else:
                coordinates = [[[]]]

            boundary = {
                'tenantId': row['Tenant ID'],
                'code': row['Boundary Code'],
                'geometry': {
                    'type': 'Polygon',
                    'coordinates': coordinates
                }
            }

            # Add description if present
            if pd.notna(row.get('Description')):
                boundary['description'] = row['Description']

            boundaries.append(boundary)

        return boundaries

    def read_boundary_relationships(self):
        """Read boundary parent-child relationships"""
        df = pd.read_excel(self.excel_file, sheet_name='Boundary_Relationships')

        if len(df) == 0:
            return []

        relationships = []
        for _, row in df.iterrows():
            relationship = {
                'tenantId': row['Tenant ID'],
                'code': row['Boundary Code'],
                'hierarchyType': row['Hierarchy Type'],
                'boundaryType': row['Boundary Type'],
                'parent': row['Parent'] if pd.notna(row['Parent']) else ''
            }
            relationships.append(relationship)

        return relationships

    def read_workflow_states(self):
        """Read workflow states"""
        df = pd.read_excel(self.excel_file, sheet_name='Workflow_States')

        if len(df) == 0:
            return []

        states = []
        for _, row in df.iterrows():
            state_code = row['State Code']

            state = {
                '_stateCode': state_code,
                'state': state_code,
                'applicationStatus': row['State Name'],
                'isStartState': str(row['Is Start State']).upper() == 'TRUE',
                'isTerminateState': str(row['Is End State']).upper() == 'TRUE',
                'actions': []
            }

            if pd.notna(row.get('SLA Days')):
                sla_days = float(row['SLA Days'])
                state['sla'] = int(sla_days * 24 * 60 * 60 * 1000)

            states.append(state)

        return states

    def read_workflow_actions(self):
        """Read workflow actions"""
        df = pd.read_excel(self.excel_file, sheet_name='Workflow_Actions')

        if len(df) == 0:
            return []

        actions = []
        for _, row in df.iterrows():
            action = {
                'fromState': row['From State'],
                'toState': row['To State'],
                'actionName': row['Action Name'],
                'roleRequired': row['Role Required'],
                'commentRequired': str(row['Comment Required']).upper() == 'TRUE'
            }
            actions.append(action)

        return actions

    def read_localization(self):
        """Read localization with auto-determination of module and locale based on code pattern"""
        try:
            df = pd.read_excel(self.excel_file, sheet_name='Localization')
        except:
            # Try lowercase sheet name
            try:
                df = pd.read_excel(self.excel_file, sheet_name='localization')
            except:
                return []

        if len(df) == 0:
            return []

        localizations = []

        for _, row in df.iterrows():
            # Skip rows with missing required fields
            if pd.notna(row.get('Code')) and pd.notna(row.get('Message')):
                code = str(row['Code']).strip()
                message = str(row['Message']).strip()

                # Determine module and locale based on code pattern
                if code.startswith('SERVICEDFS.'):
                    # Service definitions → rainmaker-pgr
                    module = 'rainmaker-pgr'
                    locale = 'en_IN'
                elif code.startswith('COMMON_MASTERS_') or code.startswith('TENANT_TENANTS_'):
                    # Common masters (departments, designations, tenants) → rainmaker-common
                    module = 'rainmaker-common'
                    locale = 'en_IN'
                else:
                    # Default fallback
                    module = 'rainmaker-common'
                    locale = 'en_IN'

                localizations.append({
                    'code': code,
                    'message': message,
                    'module': module,
                    'locale': locale
                })

        return localizations


# ============================================================================
# WORKFLOW BUILDER
# ============================================================================

def build_workflow_business_service(workflow_states, workflow_actions, config):
    """Build complete workflow BusinessService structure"""

    # Map actions to states
    for action_data in workflow_actions:
        from_state = action_data['fromState']
        to_state = action_data['toState']
        action_name = action_data['actionName']

        roles = (
            action_data['roleRequired'].split(',')
            if ',' in str(action_data['roleRequired'])
            else [action_data['roleRequired']]
        )
        roles = [role.strip() for role in roles]

        for state in workflow_states:
            state_code = state.get('_stateCode')

            if state_code == from_state:
                action_obj = {
                    'action': action_name,
                    'nextState': to_state,
                    'roles': roles,
                }

                if action_data.get('commentRequired'):
                    action_obj['isCommentRequired'] = True

                if state.get('actions') is not None:
                    state['actions'].append(action_obj)
                break

    # Clean up temporary _stateCode field
    for state in workflow_states:
        if '_stateCode' in state:
            del state['_stateCode']

    # Build final BusinessService object
    workflow_data = {
        'tenantId': config['workflow_tenant'],
        'businessService': config['workflow_business_service'],
        'business': 'pgr-services',
        'businessServiceSla': 432000000,
        'states': workflow_states,
    }

    return workflow_data


# ============================================================================
# API UPLOADER CLASS
# ============================================================================

class APIUploader:
    """Handles API uploads for PGR master data

    Service URLs are hardcoded and NOT configurable:
    - MDMS Service: :8094 (tenant, departments, designations, complaint types, employees)
    - Boundary Service: :8081 (boundaries, hierarchy, relationships)
    - Workflow Service: :8280 (workflow business services)
    - Localization Service: :8087 (localization/translations)
    """

    def __init__(self):
        # Hardcoded service URLs - DO NOT CHANGE
        self.mdms_url = "http://localhost:8094"
        self.boundary_url = "http://localhost:8081"
        self.workflow_url = "http://localhost:8280"
        self.localization_url = "http://localhost:8087"
        self.Datahandlerurl = "http://localhost:8012"

        self.auth_token = "2a57ee8c-410c-4023-a9d3-5111b4f6e304"
        self.user_info = {
            "id": 595,
            "uuid": "1fda5623-448a-4a59-ad17-657986742d67",
            "userName": "UNIFIED_DEV_USERR",
            "name": "Unified dev user",
            "mobileNumber": "8788788851",
            "emailId": "",
            "locale": None,
            "type": "EMPLOYEE",
            "roles": [
                {
                    "name": "Localisation admin",
                    "code": "LOC_ADMIN",
                    "tenantId": "pg"
                },
                {
                    "name": "Employee",
                    "code": "EMPLOYEE",
                    "tenantId": "pg"
                },
                {
                    "name": "MDMS Admin",
                    "code": "MDMS_ADMIN",
                    "tenantId": "pg"
                },
                {
                    "name": "SUPER USER",
                    "code": "SUPERUSER",
                    "tenantId": "pg"
                }
            ],
            "active": True,
            "tenantId": "pg",
            "permanentCity": None
        }
         

    def search_mdms_data(self, schema_code: str, tenant: str) -> List[Dict]:
        """Generic function to search MDMS v2 data

        Args:
            schema_code: MDMS schema code
            tenant: Tenant ID

        Returns:
            list: List of data objects retrieved
        """
        url = f"{self.mdms_url}/mdms-v2/v2/_search"

        payload = {
            "RequestInfo": {
                "apiId": "asset-services",
                "authToken": self.auth_token,
                "userInfo": self.user_info,
                "msgId": "search with from and to values"
            },
            "MdmsCriteria": {
                "tenantId": tenant,
                "schemaCode": schema_code
            }
        }

        headers = {'Content-Type': 'application/json'}

        try:
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()
            data = response.json()

            # Extract data list from response
            # API returns: {"mdms": [{"id": "...", "data": {...}, ...}]}
            mdms_records = data.get('mdms', [])
            data_list = [record['data'] for record in mdms_records]
            return data_list

        except requests.exceptions.HTTPError as e:
            print(f"HTTP Error during MDMS search for {schema_code}: {str(e)}")
            return []
        except Exception as e:
            print(f"Error during MDMS search for {schema_code}: {str(e)}")
            return []

    def create_mdms_data(self, schema_code: str, data_list: List[Dict], tenant: str, 
                        sheet_name: str = None, excel_file: str = None):
            """
            Upload MDMS data and write status directly into the uploaded Excel file
            
            Args:
                schema_code: MDMS schema code
                data_list: List of data objects to upload
                tenant: Tenant ID
                sheet_name: Excel sheet name to update with status
                excel_file: Path to the uploaded Excel file
            """
            url = f"{self.mdms_url}/mdms-v2/v2/_create/{{schema_code}}"

            results = {
                'created': 0,
                'exists': 0,
                'failed': 0,
                'errors': []
            }

            print(f"\n[UPLOADING] {schema_code}")
            print(f"   Tenant: {tenant}")
            print(f"   Records: {len(data_list)}")
            print(f"   API URL: {url}")
            print("="*60)

            # Track row-by-row status for Excel update
            row_statuses = []

            for i, data_obj in enumerate(data_list, 1):
                unique_id = (
                    data_obj.get('code') or
                    data_obj.get('serviceCode') or
                    data_obj.get('userName') or
                    str(i)
                )

                payload = {
                    "RequestInfo": {
                        "apiId": "Rainmaker",
                        "authToken": self.auth_token,
                        "userInfo": self.user_info,
                        "msgId": "1695889012604|en_IN",
                        "plainAccessRequest": {}
                    },
                    "Mdms": {
                        "tenantId": tenant,
                        "schemaCode": schema_code,
                        "uniqueIdentifier": unique_id,
                        "data": data_obj,
                        "isActive": True
                    }
                }

                headers = {'Content-Type': 'application/json'}
                status = "SUCCESS"
                status_code = 200
                error_message = ""

                try:
                    response = requests.post(url, json=payload, headers=headers)
                    status_code = response.status_code
                    response.raise_for_status()
                    print(f"   [OK] [{i}/{len(data_list)}] {unique_id}")
                    results['created'] += 1

                except requests.exceptions.HTTPError as e:
                    status_code = e.response.status_code if hasattr(e.response, 'status_code') else 500
                    error_text = e.response.text if hasattr(e.response, 'text') else str(e)
                    error_message = error_text[:200]

                    if 'already exists' in error_text.lower() or 'duplicate' in error_text.lower():
                        print(f"   [EXISTS] [{i}/{len(data_list)}] {unique_id}")
                        results['exists'] += 1
                        status = "EXISTS"
                    else:
                        print(f"   [FAILED] [{i}/{len(data_list)}] {unique_id}")
                        print(f"   ERROR: {error_message}")
                        results['failed'] += 1
                        results['errors'].append({'id': unique_id, 'error': error_message})
                        status = "FAILED"

                except Exception as e:
                    error_message = str(e)[:200]
                    status_code = 0
                    status = "FAILED"
                    print(f"   [ERROR] [{i}/{len(data_list)}] {unique_id} - {error_message[:100]}")
                    results['failed'] += 1
                    results['errors'].append({'id': unique_id, 'error': error_message})

                # Store status for this row
                row_statuses.append({
                    'row_index': i,  # 1-based index matching Excel data rows
                    'status': status,
                    'status_code': status_code,
                    'error_message': error_message
                })

                time.sleep(0.1)

            # Summary
            print("="*60)
            print(f"[SUMMARY] Created: {results['created']}")
            print(f"[SUMMARY] Already Exists: {results['exists']}")
            print(f"[SUMMARY] Failed: {results['failed']}")
            print("="*60)

            # Write status columns directly into the uploaded Excel file
            if excel_file and sheet_name and row_statuses:
                self._write_status_to_excel(
                    excel_file=excel_file,
                    sheet_name=sheet_name,
                    row_statuses=row_statuses,
                    schema_code=schema_code
                )

            return results


    



    def _write_status_to_excel(self, excel_file: str, sheet_name: str, 
                               row_statuses: List[Dict], schema_code: str):
        """
        Write / overwrite _STATUS, _STATUS_CODE, _ERROR_MESSAGE columns directly into uploaded Excel.
        - If columns exist: overwrite in-place.
        - If columns do not exist: create exactly one set of new columns at the right-most side.
        - Does NOT use ws.append() (so it won't accidentally shift or insert rows).
        Note: row_statuses[*]['row_index'] is expected to be the Excel row number you want to write to.
        If your row_index is zero-based data-row index (0..n-1) change excel_row = header_row + 1 + row_index below.
        """
        try:
            print(f"\n📝 Updating Excel file: {excel_file}")
            print(f"   Sheet: {sheet_name}")
            
            wb = load_workbook(excel_file, data_only=False)
            if sheet_name not in wb.sheetnames:
                print(f"   ⚠️  Sheet '{sheet_name}' not found - skipping status update")
                return
            ws = wb[sheet_name]
    
            # --- Find header row and map existing headers (assume header in row 1) ---
            header_row = 1
            header_map = {}
            for col in range(1, ws.max_column + 1):
                value = ws.cell(row=header_row, column=col).value
                if isinstance(value, str) and value.strip():
                    header_map[value.strip()] = col
    
            # --- Determine/create columns in a safe, non-overlapping way ---
            max_col = ws.max_column
    
            # We'll create new columns only if header missing; if we create one column,
            # we update max_col so the next created column goes to max_col+1, etc.
            def get_or_create_col(header_name):
                nonlocal max_col
                if header_name in header_map:
                    return header_map[header_name]
                else:
                    # create new column at the right
                    max_col += 1
                    header_map[header_name] = max_col
                    # set header cell (style later)
                    ws.cell(row=header_row, column=max_col, value=header_name)
                    return max_col
    
            status_col = get_or_create_col("_STATUS")
            code_col = get_or_create_col("_STATUS_CODE")
            error_col = get_or_create_col("_ERROR_MESSAGE")
    
            # --- Styles ---
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            header_font = Font(bold=True)
            success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            exists_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
            # Apply header styles (for newly created headers and to ensure consistent style)
            for header, col_idx in [("_STATUS", status_col), ("_STATUS_CODE", code_col), ("_ERROR_MESSAGE", error_col)]:
                hdr_cell = ws.cell(row=header_row, column=col_idx)
                hdr_cell.fill = header_fill
                hdr_cell.font = header_font
    
            # --- Write each row_status in-place (overwrite) ---
            for row_status in row_statuses:
                # === IMPORTANT: interpret row_index as Excel row number ===
                # If your row_status['row_index'] is zero-based relative to data rows,
                # change the next line to: excel_row = header_row + 1 + row_status['row_index']
                excel_row = row_status['row_index']
    
                # Validate excel_row is integer and >= header_row+1
                try:
                    excel_row = int(excel_row)
                except Exception:
                    # skip invalid rows
                    print(f"   ⚠️  Skipping invalid row_index: {row_status.get('row_index')}")
                    continue
                if excel_row <= header_row:
                    # If user passes header or invalid small row, shift below header
                    excel_row = header_row + 1
    
                status = row_status.get('status', '')
                status_code = row_status.get('status_code', '')
                error_msg = row_status.get('error_message', '')
    
                # Overwrite exact cells (openpyxl will create cell objects if row > current max)
                status_cell = ws.cell(row=excel_row, column=status_col, value=status)
                if status == "SUCCESS":
                    status_cell.fill = success_fill
                elif status == "EXISTS":
                    status_cell.fill = exists_fill
                elif status == "FAILED":
                    status_cell.fill = failed_fill
                else:
                    status_cell.fill = PatternFill(fill_type=None)
    
                ws.cell(row=excel_row, column=code_col, value=status_code)
                ws.cell(row=excel_row, column=error_col, value=error_msg)
    
            # --- Column widths only for newly added columns (or enforce widths always) ---
            # If you prefer to always set widths, remove the conditional checks.
            for col_idx, width in [(status_col, 15), (code_col, 15), (error_col, 50)]:
                ws.column_dimensions[get_column_letter(col_idx)].width = width
    
            # --- Protect status columns (lock cells) ---
            for r in range(1, ws.max_row + 1):
                for c in (status_col, code_col, error_col):
                    cell = ws.cell(row=r, column=c)
                    cell.protection = cell.protection.copy(locked=True)
    
            # Save
            wb.save(excel_file)
            wb.close()
            print(f"   ✅ Status columns updated successfully!")
            print(f"   📊 Updated {len(row_statuses)} rows")
    
        except Exception as e:
            print(f"   ⚠️  Could not update Excel: {str(e)}")


    def _generate_error_excel(self, failed_records: List[Dict], schema_code: str, sheet_name: str, dept_code_to_name: Dict = None) -> str:
        """Append failed records to a single consolidated error Excel file

        Args:
            failed_records: List of failed data records with status info
            schema_code: Schema code for naming
            sheet_name: Sheet name for the error file
            dept_code_to_name: Reverse mapping from department codes to names

        Returns:
            str: Path to the error Excel file
        """
        try:
            import pandas as pd
            from datetime import datetime
            import os
            from openpyxl import load_workbook

            # Create errors directory if it doesn't exist
            error_dir = 'errors'
            os.makedirs(error_dir, exist_ok=True)

            # Use a single consolidated filename
            filename = f"{error_dir}/FAILED_RECORDS.xlsx"

            # Transform records back to Excel template format
            transformed_records = self._transform_to_excel_format(failed_records, schema_code, dept_code_to_name)

            # Flatten nested structures into readable columns
            flattened_records = []
            for record in transformed_records:
                flat_record = {}

                for key, value in record.items():
                    # Handle nested objects (like city, jurisdiction, etc.)
                    if isinstance(value, dict):
                        # Flatten nested dict with prefix
                        for nested_key, nested_value in value.items():
                            flat_record[f"{key}.{nested_key}"] = nested_value
                    # Handle lists (like tenants in citymodule, roles, etc.)
                    elif isinstance(value, list):
                        if len(value) > 0:
                            # For list of dicts (like tenants: [{'code': 'pg'}, {'code': 'pg.citya'}])
                            if isinstance(value[0], dict):
                                # Extract codes/names and join with comma
                                codes = [item.get('code', item.get('name', str(item))) for item in value]
                                flat_record[key] = ', '.join(str(c) for c in codes)
                            else:
                                # Simple list (like pincodes)
                                flat_record[key] = ', '.join(str(v) for v in value)
                        else:
                            flat_record[key] = ''
                    else:
                        flat_record[key] = value

                flattened_records.append(flat_record)

            # Convert to DataFrame
            df = pd.DataFrame(flattened_records)

            # Define columns to exclude (internal/auto-generated fields only)
            exclude_cols = [
                'active', 'isActive', 'tenantId', 'uniqueIdentifier'
            ]

            # Remove excluded columns
            cols_to_keep = [col for col in df.columns if col not in exclude_cols]
            df = df[cols_to_keep]

            # Reorder columns to put status columns at the END
            status_cols = ['_STATUS', '_STATUS_CODE', '_ERROR_MESSAGE']
            other_cols = [col for col in df.columns if col not in status_cols]
            df = df[other_cols + status_cols]

            # Check if file exists
            if os.path.exists(filename):
                # Append to existing file
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # Create new file
                df.to_excel(filename, sheet_name=sheet_name, index=False)

            # Apply sheet protection to error columns
            self._protect_error_columns(filename, sheet_name, len(other_cols))

            print(f"\n📊 ERROR REPORT UPDATED:")
            print(f"   File: {filename}")
            print(f"   Sheet: {sheet_name}")
            print(f"   Failed Records: {len(failed_records)}")
            print(f"   💡 Fix the errors and re-upload this file (Error columns are protected)")

            return filename

        except Exception as e:
            print(f"\n⚠️  Could not generate error Excel: {str(e)}")
            return None

    def _transform_to_excel_format(self, records: List[Dict], schema_code: str, dept_code_to_name: Dict = None) -> List[Dict]:
        """Transform API payload format back to Excel template format

        Args:
            records: List of records in API format
            schema_code: Schema code to determine transformation rules
            dept_code_to_name: Reverse mapping from department codes to names

        Returns:
            List of records in Excel template format
        """
        if dept_code_to_name is None:
            dept_code_to_name = {}

        transformed = []

        # Handle Departments - rename 'code' and 'name' to match template
        if 'Department' in schema_code:
            for record in records:
                excel_record = {
                    'Department Name*': record.get('name', ''),
                    '_STATUS': record.get('_STATUS'),
                    '_STATUS_CODE': record.get('_STATUS_CODE'),
                    '_ERROR_MESSAGE': record.get('_ERROR_MESSAGE')
                }
                transformed.append(excel_record)

        # Handle Designations - show department name instead of code
        elif 'Designation' in schema_code:
            for record in records:
                dept_code = record.get('departmentCode', '')
                dept_name = dept_code_to_name.get(dept_code, dept_code)

                excel_record = {
                    'Department Name*': dept_name,
                    'Designation Name*': record.get('name', ''),
                    '_STATUS': record.get('_STATUS'),
                    '_STATUS_CODE': record.get('_STATUS_CODE'),
                    '_ERROR_MESSAGE': record.get('_ERROR_MESSAGE')
                }
                transformed.append(excel_record)

        # Handle Complaint Types - show department name and extract complaint type/subtype
        elif 'ServiceDefs' in schema_code:
            for record in records:
                dept_code = record.get('department', '')
                dept_name = dept_code_to_name.get(dept_code, dept_code)

                excel_record = {
                    'Complaint Type*': record.get('menuPath', ''),
                    'Complaint sub type*': record.get('name', ''),
                    'Department Name*': dept_name,
                    'Resolution Time (Hours)*': record.get('slaHours', ''),
                    'Search Words (comma separated)*': record.get('keywords', ''),
                    'Priority': record.get('priority', ''),
                    '_STATUS': record.get('_STATUS'),
                    '_STATUS_CODE': record.get('_STATUS_CODE'),
                    '_ERROR_MESSAGE': record.get('_ERROR_MESSAGE')
                }
                transformed.append(excel_record)

        # Default: return records as-is (ensuring status fields are preserved)
        else:
            # Make sure status fields are always included
            for record in records:
                excel_record = record.copy()
                # Ensure status fields exist
                if '_STATUS' not in excel_record:
                    excel_record['_STATUS'] = record.get('_STATUS')
                if '_STATUS_CODE' not in excel_record:
                    excel_record['_STATUS_CODE'] = record.get('_STATUS_CODE')
                if '_ERROR_MESSAGE' not in excel_record:
                    excel_record['_ERROR_MESSAGE'] = record.get('_ERROR_MESSAGE')
                transformed.append(excel_record)

        return transformed

    def _protect_error_columns(self, filename: str, sheet_name: str, data_col_count: int):
        """Protect the error status columns (last 3 columns) in the Excel sheet

        Args:
            filename: Path to Excel file
            sheet_name: Sheet name to protect
            data_col_count: Number of data columns (non-error columns)
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter

            wb = load_workbook(filename)
            ws = wb[sheet_name]

            # Get total columns
            total_cols = ws.max_column

            # Error columns are the last 3 columns
            error_col_start = data_col_count + 1  # +1 because Excel is 1-indexed

            # Apply gray background and bold font to error column headers
            gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            bold_font = Font(bold=True)

            for col_idx in range(error_col_start, total_cols + 1):
                col_letter = get_column_letter(col_idx)

                # Style header
                header_cell = ws[f'{col_letter}1']
                header_cell.fill = gray_fill
                header_cell.font = bold_font

                # Lock all cells in error columns (data + header)
                for row in range(1, ws.max_row + 1):
                    cell = ws[f'{col_letter}{row}']
                    cell.protection = cell.protection.copy(locked=True)

            # Unlock all data columns so users can edit them
            for col_idx in range(1, error_col_start):
                col_letter = get_column_letter(col_idx)
                for row in range(2, ws.max_row + 1):  # Skip header row
                    cell = ws[f'{col_letter}{row}']
                    cell.protection = cell.protection.copy(locked=False)

            # Protect the sheet (allow users to edit unlocked cells only)
            ws.protection.sheet = True
            ws.protection.password = ''  # Empty password (no password)
            ws.protection.selectLockedCells = True
            ws.protection.selectUnlockedCells = True

            wb.save(filename)

        except Exception as e:
            # Don't fail if protection doesn't work
            print(f"   ⚠️  Could not apply sheet protection: {str(e)}")

    def create_localization_messages(self, localization_list: List[Dict], tenant: str, sheet_name: str = 'Localization'):
        """Upload localization messages via localization service API"""
        url = f"{self.localization_url}/localization/messages/v1/_upsert"

        results = {
            'created': 0,
            'exists': 0,
            'failed': 0,
            'errors': [],
            'failed_records': []
        }

        print(f"\n[UPLOADING] Localization Messages")
        print(f"   Tenant: {tenant}")
        print(f"   Total Messages: {len(localization_list)}")
        print(f"   API URL: {url}")
        print("="*60)

        # Group messages by locale for batch upload
        from collections import defaultdict
        by_locale = defaultdict(list)
        for loc in localization_list:
            by_locale[loc['locale']].append(loc)

        print(f"\n   Found {len(by_locale)} locales: {', '.join(by_locale.keys())}")
        print("="*60)

        # Upload each locale batch
        for locale, messages in by_locale.items():
            payload = {
                "RequestInfo": {
                    "apiId": "emp",
                    "ver": "1.0",
                    "action": "create",
                    "msgId": f"{int(time.time() * 1000)}",
                    "authToken": self.auth_token,
                    "userInfo": self.user_info
                },
                "locale": locale,
                "tenantId": tenant,
                "messages": messages
            }

            headers = {'Content-Type': 'application/json'}
            status_code = None

            try:
                response = requests.post(url, json=payload, headers=headers)
                status_code = response.status_code
                response.raise_for_status()
                print(f"   [OK] Locale: {locale} - {len(messages)} messages uploaded")
                results['created'] += len(messages)

            except requests.exceptions.HTTPError as e:
                status_code = e.response.status_code if hasattr(e.response, 'status_code') else 500
                error_text = e.response.text if hasattr(e.response, 'text') else str(e)
                error_message = error_text[:400]

                # Check for duplicate/already exists errors
                if ('duplicate' in error_text.lower() or
                    'already exists' in error_text.lower() or
                    'DUPLICATE_RECORDS' in error_text or
                    'DuplicateMessageIdentityException' in error_text or
                    'unique_message_entry' in error_text.lower()):
                    print(f"   [EXISTS] Locale: {locale} - {len(messages)} messages already exist")
                    results['exists'] += len(messages)
                    # DON'T add to failed_records - already exists is not a failure!
                else:
                    # True failure
                    print(f"   [FAILED] Locale: {locale}")
                    print(f"   ERROR: {error_message}")
                    results['failed'] += len(messages)
                    results['errors'].append({
                        'locale': locale,
                        'count': len(messages),
                        'error': error_message
                    })

                    # Store failed messages for Excel export - ONLY TRUE FAILURES
                    for msg in messages:
                        failed_record = msg.copy()
                        failed_record['_STATUS'] = 'FAILED'
                        failed_record['_STATUS_CODE'] = status_code
                        failed_record['_ERROR_MESSAGE'] = error_message
                        results['failed_records'].append(failed_record)

            except Exception as e:
                error_message = str(e)[:200]
                status_code = 0
                print(f"   [ERROR] Locale: {locale} - {error_message}")
                results['failed'] += len(messages)
                results['errors'].append({
                    'locale': locale,
                    'count': len(messages),
                    'error': error_message
                })

                # Store failed messages for Excel export
                for msg in messages:
                    failed_record = msg.copy()
                    failed_record['_STATUS'] = 'FAILED'
                    failed_record['_STATUS_CODE'] = status_code
                    failed_record['_ERROR_MESSAGE'] = error_message
                    results['failed_records'].append(failed_record)

            time.sleep(0.2)

        # Summary
        print("="*60)
        print(f"[SUMMARY] Created: {results['created']}")
        print(f"[SUMMARY] Already Exists: {results['exists']}")
        print(f"[SUMMARY] Failed: {results['failed']}")

        if results['errors']:
            print(f"\n[ERRORS] Found {len(results['errors'])} error(s):")
            for err in results['errors']:
                print(f"   - Locale: {err['locale']} ({err['count']} messages)")
                print(f"     Error: {err['error'][:100]}")

        # Generate error Excel if there are failures
        if results['failed_records'] and sheet_name:
            error_file = self._generate_error_excel(results['failed_records'], 'localization.messages', sheet_name)
            results['error_file'] = error_file

        print("="*60)

        return results

    def create_boundary_hierarchy(self, hierarchy_data: Dict):
        """Create boundary hierarchy definition

        Args:
            hierarchy_data: Dict containing tenantId, hierarchyType, and boundaryHierarchy

        Returns:
            API response dict
        """
        url = f"{self.boundary_url}/boundary-service/boundary-hierarchy-definition/_create"

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": self.auth_token,
                "userInfo": self.user_info,
                "msgId": "1695889012604|en_IN",
                "plainAccessRequest": {}
            },
            'BoundaryHierarchy': hierarchy_data
        }

        headers = {'Content-Type': 'application/json'}

        try:
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()
            print(f"\n✅ [SUCCESS] Boundary hierarchy created")
            print(f"   Tenant: {hierarchy_data.get('tenantId')}")
            print(f"   Hierarchy Type: {hierarchy_data.get('hierarchyType')}")
            print(f"   Levels: {len(hierarchy_data.get('boundaryHierarchy', []))}")
            return response.json()
        except requests.exceptions.HTTPError as e:
            error_text = e.response.text if hasattr(e.response, 'text') else str(e)

            # Check if hierarchy already exists
            if 'already exists' in error_text.lower() or 'duplicate' in error_text.lower():
                print(f"\n⚠️  [EXISTS] Boundary hierarchy already exists")
                print(f"   Tenant: {hierarchy_data.get('tenantId')}")
                print(f"   Hierarchy Type: {hierarchy_data.get('hierarchyType')}")
                return {'success': True, 'exists': True}
            else:
                print(f"\n❌ [ERROR] Failed: HTTP {e.response.status_code}")
                print(f"   ERROR Details: {error_text[:500]}")
                raise
        except Exception as e:
            print(f"\n❌ [ERROR] Failed: {str(e)}")
            raise

    def create_boundary_entities(self, boundaries: List[Dict]):
        """Create boundary entities"""
        url = f"{self.boundary_url}/boundary-service/boundary/_create"

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": self.auth_token,
                "userInfo": self.user_info
            },
            'Boundary': clean_nans(boundaries)
        }

        headers = {'Content-Type': 'application/json'}

        try:
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()
            print(f"   [SUCCESS] {len(boundaries)} boundary entities created")
            return response.json()
        except requests.exceptions.HTTPError as e:
            error_text = e.response.text if hasattr(e.response, 'text') else str(e)
            print(f"   [ERROR] Failed: HTTP {e.response.status_code}")
            print(f"   ERROR Details: {error_text[:500]}")
            raise
        except Exception as e:
            print(f"   [ERROR] Failed: {str(e)}")
            raise

    def create_boundary_relationship(self, relationship: Dict):
        """Create boundary relationship"""
        url = f"{self.boundary_url}/boundary-service/boundary-relationships/_create"

        payload = {
             "RequestInfo": {
        "apiId": "asset-services",
        "ver": None,
        "ts": None,
        "action": None,
        "did": None,
        "key": None,
        "msgId": "search with from and to values",
        "authToken": "das323223-21",
        "correlationId": None,
        "userInfo": {"id":30274,"uuid":"bd5f8ea6-a022-4e74-ac2c-edf3392c6fa4","userName":"MDMSADMIN","name":"MDMSADMIN","mobileNumber":"9035169726","emailId":None,"locale":None,"type":"EMPLOYEE","roles":[{"name":"Localisation admin","code":"LOC_ADMIN","tenantId":"pg"},{"name":"MDMS Admin","code":"MDMS_ADMIN","tenantId":"pg.citya"},{"name":"MDMS Admin","code":"MDMS_ADMIN","tenantId":"pg"}],"active":True,"tenantId":"pg","permanentCity":None}
    },
            'BoundaryRelationship': relationship
        }

        headers = {'Content-Type': 'application/json'}

        try:
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()

            # Boundary relationships are processed asynchronously via Kafka
            # Wait 2 seconds to allow Kafka consumer to process and commit to DB
            time.sleep(2)

            return response.json()
        except requests.exceptions.HTTPError as e:
            error_text = e.response.text if hasattr(e.response, 'text') else str(e)
            raise Exception(f"HTTP {e.response.status_code}: {error_text}")
        except Exception as e:
            raise Exception(f"Request failed: {str(e)}")

    def create_workflow_businessservice(self, workflow_data: Dict):
        """Create workflow via Workflow-v2 API"""
        url = f"{self.workflow_url}/egov-workflow-v2/egov-wf/businessservice/_create"

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": self.auth_token,
                "userInfo": self.user_info,
                "msgId": f"{int(time.time() * 1000)}|en_IN",
                "plainAccessRequest": {}
            },
            "BusinessServices": [workflow_data]
        }

        headers = {'Content-Type': 'application/json'}

        try:
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()
            print("   [SUCCESS] Workflow created successfully")
            return response.json()

        except requests.exceptions.HTTPError as e:
            error_text = e.response.text if hasattr(e.response, 'text') else str(e)
            if 'already exists' in error_text.lower() or 'duplicate' in error_text.lower():
                print("   [EXISTS] Workflow already exists")
                return {'success': True, 'exists': True}
            else:
                print(f"   [ERROR] Failed: {error_text[:200]}")
                raise

        except Exception as e:
            print(f"   [ERROR] Failed: {str(e)}")
            raise

    def update_tenant_language(self, tenant_ids: List[str], language_label: str, language_value: str, state_tenant: str = "pg"):
        """
        Update tenant data to add a new language

        Args:
            tenant_ids: List of tenant IDs to update (e.g., ['pg.citya', 'pg.cityb'])
            language_label: Display name of language (e.g., 'हिंदी', 'ਪੰਜਾਬੀ')
            language_value: Locale code (e.g., 'hi_IN', 'pa_IN')
            state_tenant: State tenant ID (default: 'pg')

        Returns:
            Dict with results of update operations
        """
        search_url = f"{self.mdms_url}/mdms-v2/v2/_search"
        update_url = f"{self.mdms_url}/mdms-v2/v2/_update/tenant.tenants"

        results = {
            'updated': 0,
            'failed': 0,
            'skipped': 0,
            'errors': []
        }

        print(f"\n[UPDATING TENANT LANGUAGES]")
        print(f"   Adding: {language_label} ({language_value})")
        print(f"   Tenants: {', '.join(tenant_ids)}")
        print("="*60)

        for i, tenant_id in enumerate(tenant_ids, 1):
            try:
                # Step 1: Search for existing tenant data
                search_payload = {
                    "MdmsCriteria": {
                        "tenantId": state_tenant,
                        "schemaCode": "tenant.tenants",
                        "uniqueIdentifiers": [tenant_id]
                    }
                }

                headers = {'Content-Type': 'application/json'}
                search_response = requests.post(search_url, json=search_payload, headers=headers)
                search_response.raise_for_status()
                search_data = search_response.json()

                # Check if tenant exists
                if not search_data.get('mdms') or len(search_data['mdms']) == 0:
                    print(f"   [SKIP] [{i}/{len(tenant_ids)}] {tenant_id} - Tenant not found")
                    results['skipped'] += 1
                    results['errors'].append({
                        'tenant': tenant_id,
                        'error': 'Tenant not found in database'
                    })
                    continue

                mdms_record = search_data['mdms'][0]
                tenant_data = mdms_record['data']

                # Step 2: Check if language already exists
                existing_languages = tenant_data.get('languages', [])
                language_exists = any(
                    lang.get('value') == language_value
                    for lang in existing_languages
                )

                if language_exists:
                    print(f"   [EXISTS] [{i}/{len(tenant_ids)}] {tenant_id} - Language already exists")
                    results['skipped'] += 1
                    continue

                # Step 3: Add new language to the list
                new_language = {
                    "label": language_label,
                    "value": language_value
                }
                tenant_data['languages'] = existing_languages + [new_language]

                # Step 4: Update tenant with new language
                update_payload = {
                    "RequestInfo": {
                        "apiId": "asset-services",
                        "msgId": "update-language",
                        "authToken": self.auth_token,
                        "userInfo": self.user_info
                    },
                    "Mdms": {
                        "tenantId": state_tenant,
                        "schemaCode": "tenant.tenants",
                        "id": mdms_record['id'],
                        "data": tenant_data,
                        "auditDetails": mdms_record['auditDetails'],
                        "isActive": mdms_record['isActive']
                    }
                }

                update_response = requests.post(update_url, json=update_payload, headers=headers)
                update_response.raise_for_status()

                print(f"   [OK] [{i}/{len(tenant_ids)}] {tenant_id} - Language added successfully")
                results['updated'] += 1

            except requests.exceptions.HTTPError as e:
                error_text = e.response.text if hasattr(e.response, 'text') else str(e)
                print(f"   [FAILED] [{i}/{len(tenant_ids)}] {tenant_id}")
                print(f"   ERROR: {error_text[:200]}")
                results['failed'] += 1
                results['errors'].append({
                    'tenant': tenant_id,
                    'error': error_text[:200]
                })

            except Exception as e:
                print(f"   [ERROR] [{i}/{len(tenant_ids)}] {tenant_id} - {str(e)[:100]}")
                results['failed'] += 1
                results['errors'].append({
                    'tenant': tenant_id,
                    'error': str(e)[:200]
                })

            time.sleep(0.2)

        # Summary
        print("="*60)
        print(f"[SUMMARY] Updated: {results['updated']}")
        print(f"[SUMMARY] Skipped: {results['skipped']}")
        print(f"[SUMMARY] Failed: {results['failed']}")

        if results['errors']:
            print(f"\n[ERRORS] Found {len(results['errors'])} error(s):")
            for err in results['errors'][:5]:
                print(f"   - {err['tenant']}: {err['error'][:80]}")

        print("="*60)

        return results

    def setup_default_data(self, targetTenantId: str, module: str,schemaCodes: list,onlySchemas: bool = False) -> dict:

        url = f"{self.Datahandlerurl}/default-data-handler/tenant/new"

        payload = {
            "RequestInfo": {
                "apiId": "default-data-handler",
                "ver": "1.0",
                "ts": None,
                "action": "create",
                "msgId": "default-data-setup",
                "authToken": self.auth_token,
                "userInfo": self.user_info
            },
            "targetTenantId": targetTenantId
        }

        print("\n[DEFAULT DATA SETUP]")
        print(f"   Target Tenant: {targetTenantId}")
        print(f"   Module:        {module}")
        print(f"   Schemas:       {schemaCodes}")
        print(f"   Only Schemas:  {onlySchemas}")
        print(f"   API URL:       {url}")
        print("="*60)

        try:
            response = requests.post(url, json=payload, headers={"Content-Type": "application/json"})
            response.raise_for_status()

            result = response.json()

            print(f"   [SUCCESS] Default data setup complete for {targetTenantId}")
            print("="*60)

            return {
                "success": True,
                "response": result
            }

        except requests.exceptions.HTTPError as e:
            error_text = e.response.text if hasattr(e.response, "text") else str(e)
            print(f"[FAILED] HTTP ERROR {e.response.status_code}")
            print(error_text)
            print("="*60)
            return {"success": False, "error": error_text}

        except Exception as e:
            print(f"[ERROR] {str(e)}")
            print("="*60)
            return {"success": False, "error": str(e)}

    # ========================================================================
    # BOUNDARY MANAGEMENT METHODS
    # ========================================================================

    def search_boundary_hierarchies(self, tenant_id: str, limit: int = 100, offset: int = 0) -> List[Dict]:
        """Search for existing boundary hierarchies

        Args:
            tenant_id: Tenant ID to search for
            limit: Maximum number of results
            offset: Offset for pagination

        Returns:
            List of boundary hierarchy definitions
        """
        url = f"{self.boundary_url}/boundary-service/boundary-hierarchy-definition/_search"

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": self.auth_token,
                "userInfo": self.user_info,
                "msgId": "1695889012604|en_IN",
                "plainAccessRequest": {}
            },
            "BoundaryTypeHierarchySearchCriteria": {
                "tenantId": tenant_id,
                "limit": limit,
                "offset": offset
            }
        }

        headers = {'Content-Type': 'application/json'}

        try:
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()
            data = response.json()

            hierarchies = data.get('BoundaryHierarchy', [])
            print(f"\n✅ Found {len(hierarchies)} boundary hierarchy(ies) for tenant: {tenant_id}")

            for h in hierarchies:
                print(f"   • {h['hierarchyType']} ({len(h.get('boundaryHierarchy', []))} levels)")

            return hierarchies

        except requests.exceptions.HTTPError as e:
            error_text = e.response.text if hasattr(e.response, 'text') else str(e)
            print(f"❌ HTTP Error: {error_text[:200]}")
            return []
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            return []

    def generate_boundary_template(self, tenant_id: str, hierarchy_type: str, force_update: bool = True) -> Dict:
        """Generate boundary template file

        Args:
            tenant_id: Tenant ID
            hierarchy_type: Hierarchy type (e.g., 'ADMIN', 'REVENUE')
            force_update: Force regeneration

        Returns:
            Dict with generation task details
        """
        boundary_mgmt_url = "http://localhost:8099"
        url = f"{boundary_mgmt_url}/boundary-management/v1/_generate"

        params = {
            "tenantId": tenant_id,
            "hierarchyType": hierarchy_type,
            "forceUpdate": str(force_update).lower()
        }

        # Override userInfo tenantId to match the request tenant
        user_info_copy = self.user_info.copy()
        user_info_copy['tenantId'] = tenant_id

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": self.auth_token,
                "userInfo": user_info_copy,
                "msgId": f"{int(time.time() * 1000)}|en_IN",
                "plainAccessRequest": {}
            }
        }

        headers = {'Content-Type': 'application/json'}

        try:
            response = requests.post(url, json=payload, headers=headers, params=params)
            response.raise_for_status()
            data = response.json()

            resource = data.get('ResourceDetails', [{}])[0]
            print(f"\n✅ Template generation initiated")
            print(f"   Task ID: {resource.get('id')}")
            print(f"   Status: {resource.get('status')}")

            return resource

        except requests.exceptions.HTTPError as e:
            error_text = e.response.text if hasattr(e.response, 'text') else str(e)
            print(f"❌ HTTP Error: {error_text[:300]}")
            return {}
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            return {}

    def poll_boundary_template_status(self, tenant_id: str, hierarchy_type: str, max_attempts: int = 30, delay: int = 2) -> Dict:
        """Poll for boundary template generation completion

        Args:
            tenant_id: Tenant ID
            hierarchy_type: Hierarchy type
            max_attempts: Maximum polling attempts
            delay: Delay between attempts (seconds)

        Returns:
            Dict with fileStoreId when complete
        """
        boundary_mgmt_url = "http://localhost:8099"
        url = f"{boundary_mgmt_url}/boundary-management/v1/_generate-search"

        params = {
            "tenantId": tenant_id,
            "hierarchyType": hierarchy_type
        }

        # Override userInfo tenantId to match the request tenant
        user_info_copy = self.user_info.copy()
        user_info_copy['tenantId'] = tenant_id

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": self.auth_token,
                "userInfo": user_info_copy,
                "msgId": f"{int(time.time() * 1000)}|en_IN",
                "plainAccessRequest": {}
            }
        }

        headers = {'Content-Type': 'application/json'}

        print(f"\n⏳ Polling for template generation (max {max_attempts} attempts)...")

        for attempt in range(1, max_attempts + 1):
            try:
                response = requests.post(url, json=payload, headers=headers, params=params)
                response.raise_for_status()
                data = response.json()

                resources = data.get('GeneratedResource', [])
                if resources:
                    resource = resources[0]
                    status = resource.get('status')
                    filestore_id = resource.get('fileStoreid')

                    print(f"   Attempt {attempt}/{max_attempts}: Status = {status}")

                    if status == 'completed' and filestore_id:
                        print(f"\n✅ Template generation complete!")
                        print(f"   FileStore ID: {filestore_id}")
                        return resource
                    elif status == 'failed':
                        print(f"\n❌ Template generation failed")
                        return resource

                time.sleep(delay)

            except Exception as e:
                print(f"   Attempt {attempt}/{max_attempts}: Error - {str(e)[:100]}")
                time.sleep(delay)

        print(f"\n⚠️ Template generation timed out after {max_attempts} attempts")
        return {}

    def download_boundary_template(self, tenant_id: str, filestore_id: str, hierarchy_type: str = "ADMIN", output_path: str = None, return_url: bool = False):
        """Download boundary template from filestore

        Args:
            tenant_id: Tenant ID
            filestore_id: FileStore ID
            hierarchy_type: Hierarchy type for filename (optional)
            output_path: Path to save file (optional)
            return_url: If True, return dict with both path and download URL

        Returns:
            Path to downloaded file OR dict with 'path' and 'url' if return_url=True
        """
        import os
        filestore_url = "http://localhost:8009"
        url = f"{filestore_url}/filestore/v1/files/url"

        params = {
            "tenantId": tenant_id,
            "fileStoreIds": filestore_id
        }

        try:
            response = requests.get(url, params=params)
            response.raise_for_status()
            data = response.json()

            # Response format: { "fileStoreIds": [{"id": "xxx", "url": "s3://..."}] }
            file_urls = data.get('fileStoreIds', [])
            if not file_urls:
                print("❌ No file URL found in response")
                return None

            file_url = file_urls[0].get('url')
            if not file_url:
                print("❌ Invalid file URL")
                return None

            print(f"\n📥 Downloading from S3...")

            # Download the file
            file_response = requests.get(file_url)
            file_response.raise_for_status()

            # Determine output path
            if not output_path:
                os.makedirs('templates/boundary', exist_ok=True)
                output_path = f'templates/boundary/boundary_template_{tenant_id}_{hierarchy_type}.xlsx'

            with open(output_path, 'wb') as f:
                f.write(file_response.content)

            print(f"✅ Template downloaded: {output_path}")
            print(f"📊 File size: {len(file_response.content)} bytes")

            if return_url:
                return {
                    'path': output_path,
                    'url': file_url
                }
            return output_path

        except Exception as e:
            print(f"❌ Download error: {str(e)[:200]}")
            return None

    def upload_file_to_filestore(self, file_path: str, tenant_id: str, module: str = "HCM-ADMIN-CONSOLE") -> str:
        """Upload file to filestore

        Args:
            file_path: Path to file to upload
            tenant_id: Tenant ID
            module: Module name

        Returns:
            FileStore ID of uploaded file
        """
        import os
        filestore_url = "http://localhost:8009"
        url = f"{filestore_url}/filestore/v1/files"

        try:
            with open(file_path, 'rb') as f:
                files = {'file': (os.path.basename(file_path), f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                data = {
                    'tenantId': tenant_id,
                    'module': module
                }

                print(f"\n📤 Uploading file: {os.path.basename(file_path)}")
                response = requests.post(url, files=files, data=data)
                response.raise_for_status()

                result = response.json()
                files_data = result.get('files', [])

                if files_data:
                    filestore_id = files_data[0].get('fileStoreId')
                    print(f"✅ File uploaded successfully!")
                    print(f"   FileStore ID: {filestore_id}")
                    return filestore_id
                else:
                    print("❌ No filestore ID in response")
                    return None

        except Exception as e:
            print(f"❌ Upload error: {str(e)[:200]}")
            return None

    def process_boundary_data(self, tenant_id: str, filestore_id: str, hierarchy_type: str, action: str = "create") -> Dict:
        """Process uploaded boundary data

        Args:
            tenant_id: Tenant ID
            filestore_id: FileStore ID of uploaded Excel
            hierarchy_type: Hierarchy type
            action: Action type (create/update)

        Returns:
            Dict with processing results
        """
        boundary_mgmt_url = "http://localhost:8099"
        url = f"{boundary_mgmt_url}/boundary-management/v1/_process"

        # Override userInfo tenantId to match the request tenant
        user_info_copy = self.user_info.copy()
        user_info_copy['tenantId'] = tenant_id

        payload = {
            "ResourceDetails": {
                "tenantId": tenant_id,
                "fileStoreId": filestore_id,
                "hierarchyType": hierarchy_type,
                "additionalDetails": {},
                "action": action
            },
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": self.auth_token,
                "userInfo": user_info_copy,
                "msgId": f"{int(time.time() * 1000)}|en_IN",
                "plainAccessRequest": {}
            }
        }

        headers = {'Content-Type': 'application/json'}

        try:
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()
            data = response.json()

            resource = data.get('ResourceDetails', {})
            status = resource.get('status')
            processed_id = resource.get('processedFileStoreId')

            print(f"\n✅ Boundary data processing initiated")
            print(f"   Status: {status}")
            print(f"   Task ID: {resource.get('id')}")

            if processed_id:
                print(f"   Processed FileStore ID: {processed_id}")

            return resource

        except requests.exceptions.HTTPError as e:
            error_text = e.response.text if hasattr(e.response, 'text') else str(e)
            print(f"❌ HTTP Error: {error_text[:300]}")
            return {}
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            return {}
 
